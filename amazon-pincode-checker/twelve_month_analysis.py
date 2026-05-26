"""
12-Month Revenue & Blended ROAS Analysis
==========================================
Fetches:
  1. SP-API orders (365 days) → total revenue by month & category
  2. Amazon Ads API (monthly reports) → ad spend & ad sales by month
  3. Computes blended ROAS, organic %, month-over-month trends
  4. Outputs Excel for team

Usage:
  python twelve_month_analysis.py               # Full run (SP-API + Ads API)
  python twelve_month_analysis.py --skip-ads    # SP-API revenue only (faster)
  python twelve_month_analysis.py --use-cache   # Use cached data if available
"""

import os
import sys
import json
import csv
import argparse
import time
from datetime import datetime, timedelta, timezone
from collections import defaultdict

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import pandas as pd

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, LineChart, Reference
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, LineChart, Reference

from category_analysis.categories import CATEGORIES, classify_asin
from category_analysis.sp_api_client import fetch_orders_data

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
CACHE_FILE = os.path.join(SCRIPT_DIR, "category_analysis", "cache_orders_365d.csv")
ADS_CACHE_DIR = os.path.join(SCRIPT_DIR, "amazon_ads_tool", "reports", "monthly")

# ── Styles ──
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
SUBHEADER_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
BOLD = Font(name="Calibri", bold=True, size=11)
NORMAL = Font(name="Calibri", size=11)
MONEY_FMT = '₹#,##0'
PCT_FMT = '0.0"%"'
X_FMT = '0.00"x"'
THIN_BORDER = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9'),
)


def style_header(ws, row, num_cols):
    for c in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN_BORDER


def style_row(ws, row, num_cols, fill=None):
    for c in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = NORMAL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(vertical='center')
        if fill:
            cell.fill = fill


def auto_width(ws, min_w=10, max_w=40):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                val = str(cell.value or "")
                max_len = max(max_len, len(val))
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, min_w), max_w)


# ══════════════════════════════════════════════════════════════
# STEP 1: Fetch SP-API order data (365 days)
# ══════════════════════════════════════════════════════════════
def fetch_or_load_orders(use_cache=False):
    """Fetch 365 days of order data or load from cache."""
    if use_cache and os.path.exists(CACHE_FILE):
        mod_time = os.path.getmtime(CACHE_FILE)
        age_hours = (datetime.now().timestamp() - mod_time) / 3600
        if age_hours < 24:
            print(f"  Using cached 365-day order data ({age_hours:.1f}h old)")
            return pd.read_csv(CACHE_FILE)
        else:
            print(f"  Cache is {age_hours:.0f}h old, re-fetching ...")

    print("\n" + "=" * 60)
    print("  FETCHING 365 DAYS OF ORDER DATA FROM SP-API")
    print("  This will take ~10-15 minutes (12 monthly chunks)")
    print("=" * 60)

    df = fetch_orders_data(days=365)
    df.to_csv(CACHE_FILE, index=False)
    print(f"\n  Cached 365-day data to {CACHE_FILE} ({len(df)} rows)")
    return df


# ══════════════════════════════════════════════════════════════
# STEP 2: Fetch Ads data by month (Amazon Ads API)
# ══════════════════════════════════════════════════════════════
def fetch_or_load_ads_monthly(use_cache=False):
    """Fetch monthly ads data (SP campaigns + SD campaigns + SP product-level).
    
    Uses parallel submit-then-poll: submits ALL report requests first, then
    polls them all simultaneously. Much faster than sequential fetch.
    
    Applies smart date-clamping for Amazon Ads API data retention limits.
    """
    import re as _re
    os.makedirs(ADS_CACHE_DIR, exist_ok=True)

    # Ensure .env is loaded — python-dotenv may not be installed,
    # so parse the .env file manually and set env vars
    env_path = os.path.join(SCRIPT_DIR, ".env")
    if os.path.exists(env_path):
        with open(env_path) as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith('#') and '=' in line:
                    key, _, val = line.partition('=')
                    os.environ[key.strip()] = val.strip()

    from amazon_ads_tool.config import load_config as load_ads_config
    from amazon_ads_tool.api_client import AmazonAdsClient
    from amazon_ads_tool.reports import ReportManager

    ads_config = load_ads_config()
    client = AmazonAdsClient(ads_config)
    report_mgr = ReportManager(client, ads_config)

    today = datetime.now()
    monthly_ads = {}

    # Retention estimates (SP ~95 days, SD ~65 days)
    sp_retention_start = (today - timedelta(days=95)).strftime("%Y-%m-%d")
    sd_retention_start = (today - timedelta(days=65)).strftime("%Y-%m-%d")
    print(f"  Estimated retention: SP from {sp_retention_start}, SD from {sd_retention_start}")

    RETENTION = {
        "sp_campaigns": sp_retention_start,
        "sd_campaigns": sd_retention_start,
        "sp_advertised_product": sp_retention_start,
    }

    # Build list of months to fetch
    months_to_fetch = []
    for months_back in range(12, 0, -1):
        ref = today - timedelta(days=30 * months_back)
        month_start = ref.replace(day=1)
        if month_start.month == 12:
            next_month_first = month_start.replace(year=month_start.year + 1, month=1, day=1)
        else:
            next_month_first = month_start.replace(month=month_start.month + 1, day=1)
        month_end = next_month_first - timedelta(days=1)

        yesterday = today - timedelta(days=1)
        if month_end > yesterday:
            month_end = yesterday
        if month_start > yesterday:
            continue

        month_key = month_start.strftime("%Y-%m")
        months_to_fetch.append((month_key, month_start.strftime("%Y-%m-%d"), month_end.strftime("%Y-%m-%d")))

    # Also add current partial month
    cur_start = today.replace(day=1).strftime("%Y-%m-%d")
    cur_end = (today - timedelta(days=1)).strftime("%Y-%m-%d")
    cur_key = today.strftime("%Y-%m")
    if cur_start <= cur_end:
        months_to_fetch.append((cur_key, cur_start, cur_end))

    report_types = ["sp_campaigns", "sd_campaigns", "sp_advertised_product"]
    earliest_retention = min(RETENTION.values())

    # Skip months entirely before retention window
    skipped = sum(1 for _, _, end in months_to_fetch if end < earliest_retention)
    if skipped:
        print(f"  Skipping {skipped} months before retention window ({earliest_retention})")

    # ── Phase 1: Load cache + Submit all non-cached reports ──
    # pending_jobs: list of (month_key, rtype, report_id, cache_path)
    pending_jobs = []

    for month_key, start_date, end_date in months_to_fetch:
        if end_date < earliest_retention:
            monthly_ads[month_key] = {rt: [] for rt in report_types}
            continue

        month_data = {}
        all_cached = True

        for rtype in report_types:
            cache_path = os.path.join(ADS_CACHE_DIR, f"{rtype}_{month_key}.json")
            if use_cache and os.path.exists(cache_path):
                with open(cache_path) as f:
                    month_data[rtype] = json.load(f)
            else:
                all_cached = False

        if all_cached:
            print(f"    {month_key}: loaded from cache (all 3 reports)")
            monthly_ads[month_key] = month_data
            continue

        # Initialize month data with cached items
        if month_key not in monthly_ads:
            monthly_ads[month_key] = {}
        monthly_ads[month_key].update(month_data)

        print(f"    {month_key} ({start_date} to {end_date}) - submitting reports...")

        for rtype in report_types:
            cache_path = os.path.join(ADS_CACHE_DIR, f"{rtype}_{month_key}.json")
            if use_cache and os.path.exists(cache_path):
                continue  # already loaded

            # Clamp start date to retention boundary
            rt_retention = RETENTION[rtype]
            effective_start = max(start_date, rt_retention)

            if effective_start > end_date:
                print(f"      {rtype}: SKIPPED (before retention {rt_retention})")
                monthly_ads[month_key][rtype] = []
                with open(cache_path, 'w') as f:
                    json.dump([], f)
                continue

            if effective_start != start_date:
                print(f"      {rtype}: clamped {start_date} to {effective_start}")

            # Submit report request (non-blocking)
            try:
                report_id = report_mgr.request_report(rtype, start_date=effective_start, end_date=end_date)
                pending_jobs.append((month_key, rtype, report_id, cache_path))
                print(f"      {rtype}: submitted (ID: {report_id[:12]}...)")
            except Exception as e:
                err_str = str(e)
                if "400" in err_str and "retention" in err_str.lower():
                    match = _re.search(r'retention start date \((\d{4}-\d{2}-\d{2})\)', err_str)
                    if match:
                        real_retention = match.group(1)
                        RETENTION[rtype] = real_retention
                        clamped = max(start_date, real_retention)
                        if clamped <= end_date:
                            print(f"      {rtype}: retry with clamped start {clamped}")
                            try:
                                report_id = report_mgr.request_report(rtype, start_date=clamped, end_date=end_date)
                                pending_jobs.append((month_key, rtype, report_id, cache_path))
                                print(f"      {rtype}: submitted (ID: {report_id[:12]}...)")
                                continue
                            except Exception as e2:
                                print(f"      {rtype}: FAILED - {e2}")
                        else:
                            print(f"      {rtype}: SKIPPED (retention {real_retention} > end {end_date})")
                    else:
                        print(f"      {rtype}: FAILED - {e}")
                else:
                    print(f"      {rtype}: FAILED - {e}")
                monthly_ads[month_key][rtype] = []

    if not pending_jobs:
        print("  All data loaded from cache!")
        return monthly_ads

    # ── Phase 2: Poll all pending reports together ──
    print(f"\n  Waiting for {len(pending_jobs)} reports to generate on Amazon...")
    max_wait = 600
    start_time = time.time()
    completed = set()

    while len(completed) < len(pending_jobs) and (time.time() - start_time) < max_wait:
        for idx, (month_key, rtype, report_id, cache_path) in enumerate(pending_jobs):
            if idx in completed:
                continue
            try:
                result = report_mgr.client.get(
                    f"/reporting/reports/{report_id}",
                    accept="application/vnd.createasyncreportrequest.v3+json",
                )
                status = result.get("status", "")

                if status == "COMPLETED":
                    url = result.get("url", "")
                    data = report_mgr.client.download_gzip_report(url)
                    monthly_ads[month_key][rtype] = data
                    with open(cache_path, 'w') as f:
                        json.dump(data, f)
                    elapsed = int(time.time() - start_time)
                    print(f"    {month_key}/{rtype}: {len(data)} rows (took {elapsed}s)")
                    completed.add(idx)
                elif status == "FAILURE":
                    print(f"    {month_key}/{rtype}: FAILED (Amazon error)")
                    monthly_ads[month_key][rtype] = []
                    completed.add(idx)
            except Exception as e:
                print(f"    {month_key}/{rtype}: poll error - {e}")
                monthly_ads[month_key][rtype] = []
                completed.add(idx)

        if len(completed) < len(pending_jobs):
            remaining = len(pending_jobs) - len(completed)
            elapsed = int(time.time() - start_time)
            print(f"    ... {remaining} reports still pending ({elapsed}s elapsed)", flush=True)
            time.sleep(15)

    # Handle any that timed out
    for idx, (month_key, rtype, report_id, cache_path) in enumerate(pending_jobs):
        if idx not in completed:
            print(f"    {month_key}/{rtype}: TIMED OUT after {max_wait}s")
            monthly_ads[month_key][rtype] = []

    return monthly_ads


# ══════════════════════════════════════════════════════════════
# STEP 3: Process & Aggregate by Month + Category
# ══════════════════════════════════════════════════════════════
def process_orders_by_month(orders_df):
    """Break down orders into monthly totals by category."""
    df = orders_df.copy()

    # Parse dates
    df['order_date'] = pd.to_datetime(df['order_date'], errors='coerce', utc=True)
    df = df.dropna(subset=['order_date'])
    df['month'] = df['order_date'].dt.strftime('%Y-%m')
    df['item_price'] = pd.to_numeric(df['item_price'], errors='coerce').fillna(0)
    df['quantity'] = pd.to_numeric(df['quantity'], errors='coerce').fillna(1).astype(int)

    # Classify each ASIN
    df['category'] = df.apply(
        lambda r: classify_asin(str(r.get('asin', '')), str(r.get('sku', ''))),
        axis=1
    )

    # Filter to shipped/delivered only for revenue
    shipped = {'Shipped', 'Delivered', 'Shipped - Delivered to Buyer',
               'Shipped - Out for Delivery', 'Shipped - Picked Up', 'Unshipped'}
    df_rev = df[df['item_status'].isin(shipped)]

    # Monthly category totals
    monthly_cat = df_rev.groupby(['month', 'category']).agg(
        revenue=('item_price', 'sum'),
        units=('quantity', 'sum'),
        orders=('asin', 'count'),
    ).reset_index()

    # Monthly totals (all categories)
    monthly_total = df_rev.groupby('month').agg(
        revenue=('item_price', 'sum'),
        units=('quantity', 'sum'),
        orders=('asin', 'count'),
    ).reset_index()

    return monthly_cat, monthly_total


def get_ads_spend_current(ads_data):
    """Get current 30-day ads spend by category from ASIN-level data."""
    asin_ads = defaultdict(lambda: {'cost': 0, 'sales': 0})

    for p in ads_data.get('sp_products', []):
        a = p['advertisedAsin']
        asin_ads[a]['cost'] += p['cost']
        asin_ads[a]['sales'] += p.get('sales1d', 0)

    cat_ads = defaultdict(lambda: {'spend': 0, 'ad_sales': 0})
    for asin, d in asin_ads.items():
        cat = classify_asin(asin)
        cat_ads[cat]['spend'] += d['cost']
        cat_ads[cat]['ad_sales'] += d['sales']

    # Add SD
    for c in ads_data.get('sd_campaigns', []):
        # SD campaigns need campaign-name classification
        from category_analysis.categories import classify_campaign
        cat = classify_campaign(c['campaignName'])
        cat_ads[cat]['spend'] += c['cost']
        cat_ads[cat]['ad_sales'] += c.get('sales', 0)

    return dict(cat_ads)


def get_monthly_ads_by_category(monthly_ads):
    """Aggregate monthly ads data into per-month, per-category spend & sales.
    
    Returns: {month_key: {cat: {'spend': ..., 'ad_sales': ...}}}
    Also returns: {month_key: {'total_spend': ..., 'total_ad_sales': ...}}
    """
    from category_analysis.categories import classify_campaign

    monthly_cat_ads = {}
    monthly_totals_ads = {}

    for month_key, month_data in sorted(monthly_ads.items()):
        cat_ads = defaultdict(lambda: {'spend': 0, 'ad_sales': 0})

        # SP product-level data (ASIN-level, most accurate for category classification)
        for p in month_data.get('sp_advertised_product', []):
            asin = p.get('advertisedAsin', '')
            sku = p.get('advertisedSku', '')
            cat = classify_asin(asin, sku)
            cat_ads[cat]['spend'] += float(p.get('cost', 0))
            cat_ads[cat]['ad_sales'] += float(p.get('sales1d', 0))

        # SD campaign-level data (use campaign name patterns)
        for c in month_data.get('sd_campaigns', []):
            cat = classify_campaign(c.get('campaignName', ''))
            cat_ads[cat]['spend'] += float(c.get('cost', 0))
            cat_ads[cat]['ad_sales'] += float(c.get('sales', 0))

        monthly_cat_ads[month_key] = dict(cat_ads)

        total_spend = sum(d['spend'] for d in cat_ads.values())
        total_ad_sales = sum(d['ad_sales'] for d in cat_ads.values())
        monthly_totals_ads[month_key] = {'total_spend': total_spend, 'total_ad_sales': total_ad_sales}

    return monthly_cat_ads, monthly_totals_ads


# ══════════════════════════════════════════════════════════════
# STEP 4: Generate Excel
# ══════════════════════════════════════════════════════════════
def generate_excel(monthly_cat, monthly_total, cat_ads_current, monthly_cat_ads=None, monthly_totals_ads=None):
    """Create the 12-month comparison Excel."""
    wb = openpyxl.Workbook()
    months_sorted = sorted(monthly_total['month'].unique())
    has_monthly_ads = bool(monthly_cat_ads and monthly_totals_ads)

    # ── SHEET 1: Monthly Revenue Overview ──
    ws1 = wb.active
    ws1.title = "Monthly Revenue"
    ws1.sheet_properties.tabColor = "2F5496"

    ws1.merge_cells("A1:N1")
    ws1.cell(1, 1, "EMOUNT VENTURES — 12 Month Revenue & Performance").font = Font(name="Calibri", bold=True, size=14, color="2F5496")
    ws1.cell(2, 1, f"Generated: {datetime.now().strftime('%d %b %Y %H:%M')}").font = Font(name="Calibri", size=10, color="808080")

    headers = ["Month", "Total Revenue", "Units Sold", "Orders", "MoM Growth"]
    if has_monthly_ads:
        headers += ["Ad Spend", "Ad Sales", "ACoS", "Ads ROAS", "Blended ROAS"]
    r = 4
    for c, h in enumerate(headers, 1):
        ws1.cell(r, c, h)
    style_header(ws1, r, len(headers))

    r = 5
    prev_rev = 0
    total_spend_all = 0
    total_ad_sales_all = 0
    for _, row in monthly_total.sort_values('month').iterrows():
        rev = row['revenue']
        month_key = row['month']
        growth = ((rev - prev_rev) / prev_rev * 100) if prev_rev > 0 else 0

        ws1.cell(r, 1, month_key)
        ws1.cell(r, 2, round(rev)).number_format = MONEY_FMT
        ws1.cell(r, 3, int(row['units']))
        ws1.cell(r, 4, int(row['orders']))
        cell5 = ws1.cell(r, 5, round(growth, 1) if prev_rev > 0 else "-")
        if prev_rev > 0:
            cell5.number_format = '0.0"%"'

        if has_monthly_ads and month_key in monthly_totals_ads:
            m_ads = monthly_totals_ads[month_key]
            m_spend = m_ads['total_spend']
            m_ad_sales = m_ads['total_ad_sales']
            m_acos = (m_spend / m_ad_sales * 100) if m_ad_sales > 0 else 0
            m_ads_roas = (m_ad_sales / m_spend) if m_spend > 0 else 0
            m_blend_roas = (rev / m_spend) if m_spend > 0 else 0
            ws1.cell(r, 6, round(m_spend)).number_format = MONEY_FMT
            ws1.cell(r, 7, round(m_ad_sales)).number_format = MONEY_FMT
            ws1.cell(r, 8, round(m_acos, 1)).number_format = '0.0"%"'
            ws1.cell(r, 9, round(m_ads_roas, 2)).number_format = X_FMT
            ws1.cell(r, 10, round(m_blend_roas, 2)).number_format = X_FMT
            total_spend_all += m_spend
            total_ad_sales_all += m_ad_sales

        fill = GREEN_FILL if growth > 10 else (RED_FILL if growth < -10 else None)
        style_row(ws1, r, len(headers), fill=fill)
        prev_rev = rev
        r += 1

    # Totals
    r += 1
    ws1.cell(r, 1, "TOTAL (12 Months)").font = BOLD
    ws1.cell(r, 2, round(monthly_total['revenue'].sum())).number_format = MONEY_FMT
    ws1.cell(r, 3, int(monthly_total['units'].sum()))
    ws1.cell(r, 4, int(monthly_total['orders'].sum()))
    if has_monthly_ads:
        ws1.cell(r, 6, round(total_spend_all)).number_format = MONEY_FMT
        ws1.cell(r, 7, round(total_ad_sales_all)).number_format = MONEY_FMT
        tot_acos = (total_spend_all / total_ad_sales_all * 100) if total_ad_sales_all > 0 else 0
        tot_roas = (total_ad_sales_all / total_spend_all) if total_spend_all > 0 else 0
        tot_blend = (monthly_total['revenue'].sum() / total_spend_all) if total_spend_all > 0 else 0
        ws1.cell(r, 8, round(tot_acos, 1)).number_format = '0.0"%"'
        ws1.cell(r, 9, round(tot_roas, 2)).number_format = X_FMT
        ws1.cell(r, 10, round(tot_blend, 2)).number_format = X_FMT
    for c in range(1, len(headers) + 1):
        ws1.cell(r, c).font = BOLD
        ws1.cell(r, c).fill = SUBHEADER_FILL
        ws1.cell(r, c).border = THIN_BORDER

    # Add chart
    chart = BarChart()
    chart.type = "col"
    chart.title = "Monthly Revenue"
    chart.y_axis.title = "Revenue (₹)"
    chart.x_axis.title = "Month"
    chart.style = 10
    data_ref = Reference(ws1, min_col=2, min_row=4, max_row=4 + len(months_sorted))
    cats_ref = Reference(ws1, min_col=1, min_row=5, max_row=4 + len(months_sorted))
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.shape = 4
    chart.width = 25
    chart.height = 14
    ws1.add_chart(chart, f"A{r + 3}")

    auto_width(ws1)

    # ── SHEET 2: Category-wise Monthly Breakdown ──
    ws2 = wb.create_sheet("Category Monthly")
    ws2.sheet_properties.tabColor = "548235"

    ws2.cell(1, 1, "Revenue by Category — Monthly Breakdown").font = Font(name="Calibri", bold=True, size=13, color="548235")

    cat_order = ["EVA_Kids", "EVA_Gym", "ASM", "BPM", "Storage", "WTC", "UNCATEGORIZED"]
    cat_display = {k: CATEGORIES.get(k, {}).get('display_name', k) for k in cat_order}

    headers2 = ["Month"] + [cat_display.get(c, c) for c in cat_order if c in monthly_cat['category'].unique()] + ["Total"]
    active_cats = [c for c in cat_order if c in monthly_cat['category'].unique()]

    r = 3
    for c, h in enumerate(headers2, 1):
        ws2.cell(r, c, h)
    style_header(ws2, r, len(headers2))

    # Pivot data
    pivot = monthly_cat.pivot_table(index='month', columns='category', values='revenue', aggfunc='sum', fill_value=0)

    r = 4
    for month in months_sorted:
        ws2.cell(r, 1, month)
        col = 2
        total = 0
        for cat in active_cats:
            val = pivot.loc[month, cat] if month in pivot.index and cat in pivot.columns else 0
            ws2.cell(r, col, round(val)).number_format = MONEY_FMT
            total += val
            col += 1
        ws2.cell(r, col, round(total)).number_format = MONEY_FMT
        style_row(ws2, r, len(headers2))
        r += 1

    # Category totals row
    r += 1
    ws2.cell(r, 1, "12-MONTH TOTAL").font = BOLD
    col = 2
    grand = 0
    for cat in active_cats:
        val = pivot[cat].sum() if cat in pivot.columns else 0
        ws2.cell(r, col, round(val)).number_format = MONEY_FMT
        ws2.cell(r, col).font = BOLD
        ws2.cell(r, col).fill = SUBHEADER_FILL
        grand += val
        col += 1
    ws2.cell(r, col, round(grand)).number_format = MONEY_FMT
    ws2.cell(r, col).font = BOLD
    ws2.cell(r, col).fill = SUBHEADER_FILL
    ws2.cell(r, 1).fill = SUBHEADER_FILL

    # Line chart per category
    chart2 = LineChart()
    chart2.title = "Revenue Trend by Category"
    chart2.y_axis.title = "Revenue (₹)"
    chart2.style = 10
    chart2.width = 30
    chart2.height = 15
    for i, cat in enumerate(active_cats):
        data_ref = Reference(ws2, min_col=i + 2, min_row=3, max_row=3 + len(months_sorted))
        chart2.add_data(data_ref, titles_from_data=True)
    cats_ref = Reference(ws2, min_col=1, min_row=4, max_row=3 + len(months_sorted))
    chart2.set_categories(cats_ref)
    ws2.add_chart(chart2, f"A{r + 3}")

    auto_width(ws2)

    # ── SHEET 3: Monthly Blended ROAS (all months if ads data available) ──
    ws3 = wb.create_sheet("Blended ROAS")
    ws3.sheet_properties.tabColor = "BF8F00"

    if has_monthly_ads:
        ws3.cell(1, 1, "Blended ROAS — Monthly Comparison (12 Months)").font = Font(name="Calibri", bold=True, size=13, color="BF8F00")
        ws3.cell(2, 1, "Ad spend from Amazon Ads API | Total Revenue from SP-API orders").font = Font(name="Calibri", size=10, color="808080")

        headers3 = ["Month", "Total Revenue", "Ad Spend", "Ad Sales", "ACoS", "Ads ROAS",
                    "Organic Sales", "Organic %", "Blended ROAS"]
        r = 4
        for c, h in enumerate(headers3, 1):
            ws3.cell(r, c, h)
        style_header(ws3, r, len(headers3))

        rev_by_month = monthly_total.set_index('month')['revenue'].to_dict()
        r = 5
        grand_rev = 0
        grand_spend = 0
        grand_ad_sales = 0
        for month in months_sorted:
            rev = rev_by_month.get(month, 0)
            m_ads = monthly_totals_ads.get(month, {'total_spend': 0, 'total_ad_sales': 0})
            spend = m_ads['total_spend']
            ad_sales = m_ads['total_ad_sales']
            acos = (spend / ad_sales * 100) if ad_sales > 0 else 0
            ads_roas = (ad_sales / spend) if spend > 0 else 0
            organic = max(0, rev - ad_sales)
            organic_pct = (organic / rev * 100) if rev > 0 else 0
            blend_roas = (rev / spend) if spend > 0 else 0

            ws3.cell(r, 1, month)
            ws3.cell(r, 2, round(rev)).number_format = MONEY_FMT
            ws3.cell(r, 3, round(spend)).number_format = MONEY_FMT
            ws3.cell(r, 4, round(ad_sales)).number_format = MONEY_FMT
            ws3.cell(r, 5, round(acos, 1)).number_format = '0.0"%"'
            ws3.cell(r, 6, round(ads_roas, 2)).number_format = X_FMT
            ws3.cell(r, 7, round(organic)).number_format = MONEY_FMT
            ws3.cell(r, 8, round(organic_pct, 1)).number_format = '0.0"%"'
            ws3.cell(r, 9, round(blend_roas, 2)).number_format = X_FMT

            # Color code: green if blended > 12x, yellow 8-12x, red < 8x
            fill = GREEN_FILL if blend_roas >= 12 else (YELLOW_FILL if blend_roas >= 8 else (RED_FILL if spend > 0 else None))
            style_row(ws3, r, len(headers3), fill=fill)
            grand_rev += rev
            grand_spend += spend
            grand_ad_sales += ad_sales
            r += 1

        # Totals
        r += 1
        ws3.cell(r, 1, "12-MONTH TOTAL").font = BOLD
        ws3.cell(r, 2, round(grand_rev)).number_format = MONEY_FMT
        ws3.cell(r, 3, round(grand_spend)).number_format = MONEY_FMT
        ws3.cell(r, 4, round(grand_ad_sales)).number_format = MONEY_FMT
        tot_acos = (grand_spend / grand_ad_sales * 100) if grand_ad_sales > 0 else 0
        tot_roas = (grand_ad_sales / grand_spend) if grand_spend > 0 else 0
        tot_organic = max(0, grand_rev - grand_ad_sales)
        tot_organic_pct = (tot_organic / grand_rev * 100) if grand_rev > 0 else 0
        tot_blend = (grand_rev / grand_spend) if grand_spend > 0 else 0
        ws3.cell(r, 5, round(tot_acos, 1)).number_format = '0.0"%"'
        ws3.cell(r, 6, round(tot_roas, 2)).number_format = X_FMT
        ws3.cell(r, 7, round(tot_organic)).number_format = MONEY_FMT
        ws3.cell(r, 8, round(tot_organic_pct, 1)).number_format = '0.0"%"'
        ws3.cell(r, 9, round(tot_blend, 2)).number_format = X_FMT
        for c in range(1, len(headers3) + 1):
            ws3.cell(r, c).font = BOLD
            ws3.cell(r, c).fill = SUBHEADER_FILL
            ws3.cell(r, c).border = THIN_BORDER

        # Blended ROAS trend chart
        chart3 = LineChart()
        chart3.title = "Blended ROAS Trend (Monthly)"
        chart3.y_axis.title = "ROAS (x)"
        chart3.style = 10
        chart3.width = 25
        chart3.height = 14
        blend_ref = Reference(ws3, min_col=9, min_row=4, max_row=4 + len(months_sorted))
        ads_roas_ref = Reference(ws3, min_col=6, min_row=4, max_row=4 + len(months_sorted))
        chart3.add_data(blend_ref, titles_from_data=True)
        chart3.add_data(ads_roas_ref, titles_from_data=True)
        cats_ref3 = Reference(ws3, min_col=1, min_row=5, max_row=4 + len(months_sorted))
        chart3.set_categories(cats_ref3)
        ws3.add_chart(chart3, f"A{r + 3}")

    else:
        # Fallback: current-period-only blended ROAS (old behavior)
        ws3.cell(1, 1, "Blended ROAS — Current Period (Last 30 Days)").font = Font(name="Calibri", bold=True, size=13, color="BF8F00")
        ws3.cell(2, 1, "Ad spend from Amazon Ads API | Total Revenue from SP-API orders").font = Font(name="Calibri", size=10, color="808080")

        headers3 = ["Category", "Ad Spend", "Ad Sales", "Ads ROAS", "Target Ads ROI",
                    "Total Revenue", "Organic Sales", "Organic %", "Blended ROAS", "Target Blended", "Status"]
        r = 4
        for c, h in enumerate(headers3, 1):
            ws3.cell(r, c, h)
        style_header(ws3, r, len(headers3))

        if months_sorted:
            latest_month = months_sorted[-1]
            last_month_rev = monthly_cat[monthly_cat['month'] == latest_month].set_index('category')['revenue'].to_dict()
        else:
            last_month_rev = {}

        r = 5
        totals = {'spend': 0, 'ad_sales': 0, 'revenue': 0}
        for cat_key in ["EVA_Kids", "EVA_Gym", "ASM", "BPM", "Storage"]:
            cat_cfg = CATEGORIES.get(cat_key, {})
            disp = cat_cfg.get('display_name', cat_key)
            t_ads = cat_cfg.get('target_ads_roi', 0)
            t_blend = cat_cfg.get('target_blended_roi', 0)

            ads = cat_ads_current.get(cat_key, {'spend': 0, 'ad_sales': 0})
            spend = ads['spend']
            ad_sales = ads['ad_sales']
            revenue = last_month_rev.get(cat_key, 0)
            organic = max(0, revenue - ad_sales) if revenue > 0 else 0
            organic_pct = (organic / revenue * 100) if revenue > 0 else 0
            ads_roas = (ad_sales / spend) if spend > 0 else 0
            blend_roas = (revenue / spend) if spend > 0 and revenue > 0 else 0

            if t_ads > 0:
                if ads_roas >= t_ads:
                    status = "ON TARGET"
                    fill = GREEN_FILL
                elif ads_roas >= t_ads * 0.8:
                    status = "CLOSE"
                    fill = YELLOW_FILL
                else:
                    status = "BELOW"
                    fill = RED_FILL
            else:
                status = "-"
                fill = None

            ws3.cell(r, 1, disp).font = BOLD
            ws3.cell(r, 2, round(spend)).number_format = MONEY_FMT
            ws3.cell(r, 3, round(ad_sales)).number_format = MONEY_FMT
            ws3.cell(r, 4, round(ads_roas, 2)).number_format = X_FMT
            ws3.cell(r, 5, f"{t_ads}x")
            ws3.cell(r, 6, round(revenue)).number_format = MONEY_FMT
            ws3.cell(r, 7, round(organic)).number_format = MONEY_FMT
            ws3.cell(r, 8, round(organic_pct, 1)).number_format = PCT_FMT
            ws3.cell(r, 9, round(blend_roas, 2)).number_format = X_FMT
            ws3.cell(r, 10, f"{t_blend}x")
            ws3.cell(r, 11, status)
            style_row(ws3, r, len(headers3), fill=fill)

            totals['spend'] += spend
            totals['ad_sales'] += ad_sales
            totals['revenue'] += revenue
            r += 1

        r += 1
        ws3.cell(r, 1, "PORTFOLIO TOTAL").font = Font(name="Calibri", bold=True, size=12)
        ws3.cell(r, 2, round(totals['spend'])).number_format = MONEY_FMT
        ws3.cell(r, 3, round(totals['ad_sales'])).number_format = MONEY_FMT
        t_roas = (totals['ad_sales'] / totals['spend']) if totals['spend'] > 0 else 0
        t_blend_val = (totals['revenue'] / totals['spend']) if totals['spend'] > 0 and totals['revenue'] > 0 else 0
        t_organic = max(0, totals['revenue'] - totals['ad_sales'])
        t_organic_pct = (t_organic / totals['revenue'] * 100) if totals['revenue'] > 0 else 0
        ws3.cell(r, 4, round(t_roas, 2)).number_format = X_FMT
        ws3.cell(r, 6, round(totals['revenue'])).number_format = MONEY_FMT
        ws3.cell(r, 7, round(t_organic)).number_format = MONEY_FMT
        ws3.cell(r, 8, round(t_organic_pct, 1)).number_format = PCT_FMT
        ws3.cell(r, 9, round(t_blend_val, 2)).number_format = X_FMT
        for c in range(1, len(headers3) + 1):
            ws3.cell(r, c).font = BOLD
            ws3.cell(r, c).fill = SUBHEADER_FILL
            ws3.cell(r, c).border = THIN_BORDER

    auto_width(ws3)

    # ── SHEET 4: Monthly Ads by Category (only when monthly ads data available) ──
    if has_monthly_ads:
        ws_ads = wb.create_sheet("Monthly Ads by Category")
        ws_ads.sheet_properties.tabColor = "C00000"

        ws_ads.cell(1, 1, "Monthly Ad Spend & ROAS by Category").font = Font(name="Calibri", bold=True, size=13, color="C00000")
        ws_ads.cell(2, 1, "Source: Amazon Ads API (SP + SD)").font = Font(name="Calibri", size=10, color="808080")

        main_cats = ["EVA_Kids", "EVA_Gym", "ASM", "BPM", "Storage"]
        rev_by_month_cat = monthly_cat.set_index(['month', 'category'])['revenue'].to_dict()

        # For each category, write a section
        r = 4
        for cat_key in main_cats:
            cat_cfg = CATEGORIES.get(cat_key, {})
            disp = cat_cfg.get('display_name', cat_key)
            t_ads = cat_cfg.get('target_ads_roi', 0)
            t_blend = cat_cfg.get('target_blended_roi', 0)

            ws_ads.cell(r, 1, f"{disp}").font = Font(name="Calibri", bold=True, size=12, color="C00000")
            ws_ads.cell(r, 2, f"Target Ads ROAS: {t_ads}x | Target Blended: {t_blend}x").font = Font(name="Calibri", size=10, color="808080")
            r += 1

            cat_headers = ["Month", "Ad Spend", "Ad Sales", "ACoS", "Ads ROAS",
                          "Total Revenue", "Blended ROAS", "Status"]
            for c, h in enumerate(cat_headers, 1):
                ws_ads.cell(r, c, h)
            style_header(ws_ads, r, len(cat_headers))
            r += 1

            cat_total_spend = 0
            cat_total_ad_sales = 0
            cat_total_rev = 0

            for month in months_sorted:
                m_cat_ads = monthly_cat_ads.get(month, {}).get(cat_key, {'spend': 0, 'ad_sales': 0})
                spend = m_cat_ads['spend']
                ad_sales = m_cat_ads['ad_sales']
                rev = rev_by_month_cat.get((month, cat_key), 0)
                acos = (spend / ad_sales * 100) if ad_sales > 0 else 0
                ads_roas = (ad_sales / spend) if spend > 0 else 0
                blend_roas = (rev / spend) if spend > 0 else 0

                if t_ads > 0 and spend > 0:
                    if ads_roas >= t_ads:
                        status = "ON TARGET"
                        fill = GREEN_FILL
                    elif ads_roas >= t_ads * 0.8:
                        status = "CLOSE"
                        fill = YELLOW_FILL
                    else:
                        status = "BELOW"
                        fill = RED_FILL
                else:
                    status = "-"
                    fill = None

                ws_ads.cell(r, 1, month)
                ws_ads.cell(r, 2, round(spend)).number_format = MONEY_FMT
                ws_ads.cell(r, 3, round(ad_sales)).number_format = MONEY_FMT
                ws_ads.cell(r, 4, round(acos, 1)).number_format = '0.0"%"'
                ws_ads.cell(r, 5, round(ads_roas, 2)).number_format = X_FMT
                ws_ads.cell(r, 6, round(rev)).number_format = MONEY_FMT
                ws_ads.cell(r, 7, round(blend_roas, 2)).number_format = X_FMT
                ws_ads.cell(r, 8, status)
                style_row(ws_ads, r, len(cat_headers), fill=fill)

                cat_total_spend += spend
                cat_total_ad_sales += ad_sales
                cat_total_rev += rev
                r += 1

            # Category total
            ws_ads.cell(r, 1, "TOTAL").font = BOLD
            ws_ads.cell(r, 2, round(cat_total_spend)).number_format = MONEY_FMT
            ws_ads.cell(r, 3, round(cat_total_ad_sales)).number_format = MONEY_FMT
            c_acos = (cat_total_spend / cat_total_ad_sales * 100) if cat_total_ad_sales > 0 else 0
            c_roas = (cat_total_ad_sales / cat_total_spend) if cat_total_spend > 0 else 0
            c_blend = (cat_total_rev / cat_total_spend) if cat_total_spend > 0 else 0
            ws_ads.cell(r, 4, round(c_acos, 1)).number_format = '0.0"%"'
            ws_ads.cell(r, 5, round(c_roas, 2)).number_format = X_FMT
            ws_ads.cell(r, 6, round(cat_total_rev)).number_format = MONEY_FMT
            ws_ads.cell(r, 7, round(c_blend, 2)).number_format = X_FMT
            for c in range(1, len(cat_headers) + 1):
                ws_ads.cell(r, c).font = BOLD
                ws_ads.cell(r, c).fill = SUBHEADER_FILL
                ws_ads.cell(r, c).border = THIN_BORDER
            r += 2  # Gap before next category

        auto_width(ws_ads)

    # ── SHEET 4: Category Monthly Units ──
    ws4 = wb.create_sheet("Units by Category")
    ws4.sheet_properties.tabColor = "7030A0"

    ws4.cell(1, 1, "Units Sold by Category — Monthly").font = Font(name="Calibri", bold=True, size=13, color="7030A0")

    headers4 = ["Month"] + [cat_display.get(c, c) for c in active_cats] + ["Total"]
    r = 3
    for c, h in enumerate(headers4, 1):
        ws4.cell(r, c, h)
    style_header(ws4, r, len(headers4))

    pivot_units = monthly_cat.pivot_table(index='month', columns='category', values='units', aggfunc='sum', fill_value=0)
    r = 4
    for month in months_sorted:
        ws4.cell(r, 1, month)
        col = 2
        total = 0
        for cat in active_cats:
            val = int(pivot_units.loc[month, cat]) if month in pivot_units.index and cat in pivot_units.columns else 0
            ws4.cell(r, col, val)
            total += val
            col += 1
        ws4.cell(r, col, total)
        style_row(ws4, r, len(headers4))
        r += 1

    auto_width(ws4)

    # Save
    ts = datetime.now().strftime("%Y%m%d_%H%M")
    out_dir = os.path.join(SCRIPT_DIR, "output")
    os.makedirs(out_dir, exist_ok=True)
    out_path = os.path.join(out_dir, f"12_Month_Performance_{ts}.xlsx")
    wb.save(out_path)
    print(f"\n  Excel saved to: {out_path}")
    return out_path


# ══════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════
def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--skip-ads", action="store_true", help="Skip ads API fetch")
    parser.add_argument("--use-cache", action="store_true", help="Use cached data if available")
    args = parser.parse_args()

    print("\n" + "=" * 60)
    print("  12-MONTH REVENUE & BLENDED ROAS ANALYSIS")
    print("=" * 60)

    # Step 1: Orders data
    print("\n[1/3] Fetching SP-API order data (365 days) ...")
    orders_df = fetch_or_load_orders(use_cache=args.use_cache)
    print(f"  Total rows: {len(orders_df)}")

    # Step 2: Ads data
    print("\n[2/3] Loading ads data ...")
    monthly_ads = None
    monthly_cat_ads = None
    monthly_totals_ads = None
    cat_ads_current = {}

    if args.skip_ads:
        # Just load current 30-day snapshot for basic reporting
        ads_data = {}
        prod_file = os.path.join(SCRIPT_DIR, "amazon_ads_tool", "reports", "sp_advertised_product_data.json")
        sd_file = os.path.join(SCRIPT_DIR, "amazon_ads_tool", "reports", "sd_campaigns_data.json")
        if os.path.exists(prod_file):
            with open(prod_file) as f:
                ads_data['sp_products'] = json.load(f)
        if os.path.exists(sd_file):
            with open(sd_file) as f:
                ads_data['sd_campaigns'] = json.load(f)
        cat_ads_current = get_ads_spend_current(ads_data)
    else:
        # Fetch monthly ads reports from Ads API
        monthly_ads = fetch_or_load_ads_monthly(use_cache=args.use_cache)
        monthly_cat_ads, monthly_totals_ads = get_monthly_ads_by_category(monthly_ads)

        # Print monthly ads summary
        print("\n  " + "─" * 50)
        print("  MONTHLY ADS SUMMARY")
        print("  " + "─" * 50)
        for mk in sorted(monthly_totals_ads.keys()):
            mt = monthly_totals_ads[mk]
            roas = (mt['total_ad_sales'] / mt['total_spend']) if mt['total_spend'] > 0 else 0
            print(f"  {mk}:  Spend ₹{mt['total_spend']:>8,.0f}  |  Ad Sales ₹{mt['total_ad_sales']:>9,.0f}  |  ROAS {roas:.2f}x")
        print("  " + "─" * 50)

    # Step 3: Process
    print("\n[3/3] Processing monthly breakdown ...")
    monthly_cat, monthly_total = process_orders_by_month(orders_df)
    print(f"  Months found: {sorted(monthly_total['month'].unique())}")

    # Quick summary
    print("\n" + "─" * 60)
    print("  MONTHLY REVENUE SUMMARY")
    print("─" * 60)
    for _, row in monthly_total.sort_values('month').iterrows():
        print(f"  {row['month']}:  ₹{row['revenue']:>10,.0f}  |  {int(row['units']):>5} units  |  {int(row['orders']):>5} orders")
    print(f"\n  12-Month Total: ₹{monthly_total['revenue'].sum():,.0f}")
    print("─" * 60)

    # Generate Excel
    out_path = generate_excel(monthly_cat, monthly_total, cat_ads_current,
                              monthly_cat_ads=monthly_cat_ads,
                              monthly_totals_ads=monthly_totals_ads)

    print("\n  Done!")
    return out_path


if __name__ == "__main__":
    main()
