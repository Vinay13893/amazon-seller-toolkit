"""Generate Excel report for team from category analysis data."""
import json
import os
import sys
from collections import defaultdict
from datetime import datetime

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Installing openpyxl...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    from openpyxl.utils import get_column_letter

from category_analysis.categories import CATEGORIES, classify_asin

# ── Paths ──
ADS_DIR = os.path.join(os.path.dirname(__file__), "amazon_ads_tool", "reports")
sp_file = os.path.join(ADS_DIR, "sp_campaigns_data.json")
sd_file = os.path.join(ADS_DIR, "sd_campaigns_data.json")
tgt_file = os.path.join(ADS_DIR, "sp_targeting_data.json")
st_file = os.path.join(ADS_DIR, "sp_search_terms_data.json")
prod_file = os.path.join(ADS_DIR, "sp_advertised_product_data.json")

# ── Load data ──
with open(sp_file) as f: sp_data = json.load(f)
with open(sd_file) as f: sd_data = json.load(f)
with open(tgt_file) as f: tgt_data = json.load(f)
with open(st_file) as f: st_data = json.load(f)
with open(prod_file) as f: prod_data = json.load(f)

# ── Load SP-API order data if available ──
cache_csv = os.path.join(os.path.dirname(__file__), "category_analysis", "cache_orders_30d.csv")
asin_revenue = {}
if os.path.exists(cache_csv):
    import csv
    with open(cache_csv, newline="", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            asin = row.get("asin", "").strip().upper()
            rev = float(row.get("total_revenue", 0) or 0)
            if asin:
                asin_revenue[asin] = asin_revenue.get(asin, 0) + rev

# ── Styles ──
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
SUBHEADER_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
BOLD = Font(name="Calibri", bold=True, size=11)
NORMAL = Font(name="Calibri", size=11)
MONEY_FMT = '₹#,##0'
MONEY_DEC = '₹#,##0.00'
PCT_FMT = '0.0"%"'
X_FMT = '0.00"x"'
THIN_BORDER = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9'),
)

def style_header(ws, row, num_cols):
    for c in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN_BORDER

def style_row(ws, row, num_cols, fill=None):
    for c in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = NORMAL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(vertical='center')
        if fill:
            cell.fill = fill

def auto_width(ws, min_w=10, max_w=40):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                val = str(cell.value or "")
                max_len = max(max_len, len(val))
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, min_w), max_w)


wb = openpyxl.Workbook()

# ══════════════════════════════════════════════════════════════
# SHEET 1: Category Summary (Targets vs Actual)
# ══════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Category Summary"
ws1.sheet_properties.tabColor = "2F5496"

# Title
ws1.merge_cells("A1:L1")
ws1.cell(1, 1, "EMOUNT VENTURES — 30-Day Category Performance Report").font = Font(name="Calibri", bold=True, size=14, color="2F5496")
ws1.cell(2, 1, f"Generated: {datetime.now().strftime('%d %b %Y %H:%M')}").font = Font(name="Calibri", size=10, color="808080")
ws1.row_dimensions[1].height = 30

# Aggregate ASIN-level ad data by category
asin_ads = defaultdict(lambda: {'cost': 0, 'sales': 0, 'clicks': 0, 'imps': 0, 'orders': 0, 'units': 0, 'sku': ''})
for p in prod_data:
    a = p['advertisedAsin']
    asin_ads[a]['cost'] += p['cost']
    asin_ads[a]['sales'] += p.get('sales1d', 0)
    asin_ads[a]['clicks'] += p['clicks']
    asin_ads[a]['imps'] += p['impressions']
    asin_ads[a]['orders'] += p.get('purchases1d', 0)
    asin_ads[a]['units'] += p.get('unitsSoldClicks1d', 0)
    if p.get('advertisedSku'):
        asin_ads[a]['sku'] = p['advertisedSku']

cat_summary = {}
for asin, d in asin_ads.items():
    cat = classify_asin(asin, d['sku'])
    if cat not in cat_summary:
        cat_summary[cat] = {'spend': 0, 'ad_sales': 0, 'orders': 0, 'units': 0, 'clicks': 0, 'imps': 0, 'total_rev': 0, 'asins': 0}
    cat_summary[cat]['spend'] += d['cost']
    cat_summary[cat]['ad_sales'] += d['sales']
    cat_summary[cat]['orders'] += d['orders']
    cat_summary[cat]['units'] += d['units']
    cat_summary[cat]['clicks'] += d['clicks']
    cat_summary[cat]['imps'] += d['imps']
    cat_summary[cat]['total_rev'] += asin_revenue.get(asin, 0)
    cat_summary[cat]['asins'] += 1

headers = ["Category", "ASINs", "Ad Spend", "Ad Sales", "Orders", "ACoS", "Target ACoS",
           "Ads ROAS", "Target Ads ROI", "Total Revenue", "Blended ROAS", "Target Blended", "Status"]
r = 4
for c, h in enumerate(headers, 1):
    ws1.cell(r, c, h)
style_header(ws1, r, len(headers))

cat_order = ["EVA_Kids", "EVA_Gym", "ASM", "BPM", "Storage", "WTC", "UNCATEGORIZED"]
r = 5
portfolio = {'spend': 0, 'ad_sales': 0, 'orders': 0, 'total_rev': 0}

for cat_key in cat_order:
    if cat_key not in cat_summary:
        continue
    d = cat_summary[cat_key]
    cat_cfg = CATEGORIES.get(cat_key, {})
    disp = cat_cfg.get('display_name', cat_key)
    t_acos = cat_cfg.get('target_acos', '-')
    t_ads = cat_cfg.get('target_ads_roi', '-')
    t_blend = cat_cfg.get('target_blended_roi', '-')
    acos = (d['spend'] / d['ad_sales'] * 100) if d['ad_sales'] > 0 else 0
    roas = (d['ad_sales'] / d['spend']) if d['spend'] > 0 else 0
    blend = (d['total_rev'] / d['spend']) if d['spend'] > 0 and d['total_rev'] > 0 else 0

    # Status
    if isinstance(t_ads, (int, float)) and t_ads > 0:
        if roas >= t_ads:
            status = "✓ ON TARGET"
            fill = GREEN_FILL
        elif roas >= t_ads * 0.8:
            status = "~ CLOSE"
            fill = YELLOW_FILL
        else:
            status = "✗ BELOW TARGET"
            fill = RED_FILL
    else:
        status = "-"
        fill = None

    ws1.cell(r, 1, disp)
    ws1.cell(r, 2, d['asins'])
    ws1.cell(r, 3, round(d['spend'])).number_format = MONEY_FMT
    ws1.cell(r, 4, round(d['ad_sales'])).number_format = MONEY_FMT
    ws1.cell(r, 5, d['orders'])
    ws1.cell(r, 6, round(acos, 1)).number_format = '0.0"%"'
    ws1.cell(r, 7, t_acos if isinstance(t_acos, str) else f"{t_acos}%")
    ws1.cell(r, 8, round(roas, 2)).number_format = '0.00"x"'
    ws1.cell(r, 9, f"{t_ads}x" if isinstance(t_ads, (int, float)) else t_ads)
    ws1.cell(r, 10, round(d['total_rev']) if d['total_rev'] else "N/A").number_format = MONEY_FMT
    ws1.cell(r, 11, round(blend, 2) if blend else "N/A").number_format = '0.00"x"'
    ws1.cell(r, 12, f"{t_blend}x" if isinstance(t_blend, (int, float)) else t_blend)
    ws1.cell(r, 13, status)

    style_row(ws1, r, len(headers), fill=fill)
    ws1.cell(r, 1).font = BOLD

    portfolio['spend'] += d['spend']
    portfolio['ad_sales'] += d['ad_sales']
    portfolio['orders'] += d['orders']
    portfolio['total_rev'] += d['total_rev']
    r += 1

# Portfolio total row
r += 1
ws1.cell(r, 1, "PORTFOLIO TOTAL").font = Font(name="Calibri", bold=True, size=12)
ws1.cell(r, 3, round(portfolio['spend'])).number_format = MONEY_FMT
ws1.cell(r, 4, round(portfolio['ad_sales'])).number_format = MONEY_FMT
ws1.cell(r, 5, portfolio['orders'])
p_acos = (portfolio['spend'] / portfolio['ad_sales'] * 100) if portfolio['ad_sales'] > 0 else 0
p_roas = (portfolio['ad_sales'] / portfolio['spend']) if portfolio['spend'] > 0 else 0
p_blend = (portfolio['total_rev'] / portfolio['spend']) if portfolio['spend'] > 0 and portfolio['total_rev'] > 0 else 0
ws1.cell(r, 6, round(p_acos, 1)).number_format = '0.0"%"'
ws1.cell(r, 8, round(p_roas, 2)).number_format = '0.00"x"'
ws1.cell(r, 10, round(portfolio['total_rev']) if portfolio['total_rev'] else "N/A").number_format = MONEY_FMT
ws1.cell(r, 11, round(p_blend, 2) if p_blend else "N/A").number_format = '0.00"x"'
for c in range(1, len(headers) + 1):
    ws1.cell(r, c).border = THIN_BORDER
    ws1.cell(r, c).font = BOLD
    ws1.cell(r, c).fill = SUBHEADER_FILL

auto_width(ws1)

# ══════════════════════════════════════════════════════════════
# SHEET 2: Top SP Campaigns
# ══════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("SP Campaigns")
ws2.sheet_properties.tabColor = "548235"

headers2 = ["Campaign", "Status", "Budget", "Spend", "Sales", "Orders", "ACoS", "ROAS", "Clicks", "Impressions", "CTR", "CPC", "CVR"]
ws2.cell(1, 1, "SP Campaign Performance — Last 30 Days").font = Font(name="Calibri", bold=True, size=13, color="548235")
r = 3
for c, h in enumerate(headers2, 1):
    ws2.cell(r, c, h)
style_header(ws2, r, len(headers2))

sp_sorted = sorted(sp_data, key=lambda x: x.get('sales1d', 0), reverse=True)
r = 4
for camp in sp_sorted:
    s = camp.get('sales1d', 0)
    co = camp['cost']
    clicks = camp['clicks']
    imps = camp['impressions']
    orders = camp.get('purchases1d', 0)
    acos = (co / s * 100) if s > 0 else 0
    roas = (s / co) if co > 0 else 0
    ctr = (clicks / imps * 100) if imps > 0 else 0
    cpc = (co / clicks) if clicks > 0 else 0
    cvr = (orders / clicks * 100) if clicks > 0 else 0

    ws2.cell(r, 1, camp['campaignName'])
    ws2.cell(r, 2, camp.get('campaignStatus', ''))
    ws2.cell(r, 3, camp.get('campaignBudgetAmount', 0)).number_format = MONEY_FMT
    ws2.cell(r, 4, round(co)).number_format = MONEY_FMT
    ws2.cell(r, 5, round(s)).number_format = MONEY_FMT
    ws2.cell(r, 6, orders)
    ws2.cell(r, 7, round(acos, 1)).number_format = '0.0"%"'
    ws2.cell(r, 8, round(roas, 2)).number_format = '0.00"x"'
    ws2.cell(r, 9, clicks)
    ws2.cell(r, 10, imps)
    ws2.cell(r, 11, round(ctr, 2)).number_format = '0.00"%"'
    ws2.cell(r, 12, round(cpc, 2)).number_format = MONEY_DEC
    ws2.cell(r, 13, round(cvr, 2)).number_format = '0.00"%"'

    # Color coding
    if s == 0 and co > 200:
        fill = RED_FILL
    elif acos > 30 and co > 500:
        fill = YELLOW_FILL
    else:
        fill = None
    style_row(ws2, r, len(headers2), fill=fill)
    r += 1

auto_width(ws2)

# ══════════════════════════════════════════════════════════════
# SHEET 3: ASIN Performance
# ══════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("ASIN Performance")
ws3.sheet_properties.tabColor = "BF8F00"

headers3 = ["ASIN", "SKU", "Category", "Spend", "Ad Sales", "Total Revenue",
            "Organic Sales", "Orders", "ACoS", "Ads ROAS", "Blended ROAS", "Clicks", "CVR"]
ws3.cell(1, 1, "ASIN-Level Performance — Last 30 Days").font = Font(name="Calibri", bold=True, size=13, color="BF8F00")
r = 3
for c, h in enumerate(headers3, 1):
    ws3.cell(r, c, h)
style_header(ws3, r, len(headers3))

asin_list = sorted(asin_ads.items(), key=lambda x: x[1]['sales'], reverse=True)
r = 4
for asin, d in asin_list:
    cat = classify_asin(asin, d['sku'])
    cat_disp = CATEGORIES.get(cat, {}).get('display_name', cat)
    total_rev = asin_revenue.get(asin, 0)
    organic = max(0, total_rev - d['sales']) if total_rev > 0 else 0
    acos = (d['cost'] / d['sales'] * 100) if d['sales'] > 0 else 0
    roas = (d['sales'] / d['cost']) if d['cost'] > 0 else 0
    blend = (total_rev / d['cost']) if d['cost'] > 0 and total_rev > 0 else 0
    cvr = (d['orders'] / d['clicks'] * 100) if d['clicks'] > 0 else 0

    ws3.cell(r, 1, asin)
    ws3.cell(r, 2, d['sku'])
    ws3.cell(r, 3, cat_disp)
    ws3.cell(r, 4, round(d['cost'])).number_format = MONEY_FMT
    ws3.cell(r, 5, round(d['sales'])).number_format = MONEY_FMT
    ws3.cell(r, 6, round(total_rev) if total_rev else "N/A").number_format = MONEY_FMT
    ws3.cell(r, 7, round(organic) if total_rev else "N/A").number_format = MONEY_FMT
    ws3.cell(r, 8, d['orders'])
    ws3.cell(r, 9, round(acos, 1)).number_format = '0.0"%"'
    ws3.cell(r, 10, round(roas, 2)).number_format = '0.00"x"'
    ws3.cell(r, 11, round(blend, 2) if blend else "N/A").number_format = '0.00"x"'
    ws3.cell(r, 12, d['clicks'])
    ws3.cell(r, 13, round(cvr, 1)).number_format = '0.0"%"'

    if d['sales'] == 0 and d['cost'] > 100:
        fill = RED_FILL
    else:
        fill = None
    style_row(ws3, r, len(headers3), fill=fill)
    r += 1

auto_width(ws3)

# ══════════════════════════════════════════════════════════════
# SHEET 4: Top Search Terms
# ══════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Search Terms")
ws4.sheet_properties.tabColor = "7030A0"

headers4 = ["Search Term", "Campaign", "Targeting", "Clicks", "Spend", "Sales", "Orders", "ACoS", "ROAS"]
ws4.cell(1, 1, "Top Converting Search Terms — Last 30 Days").font = Font(name="Calibri", bold=True, size=13, color="7030A0")
r = 3
for c, h in enumerate(headers4, 1):
    ws4.cell(r, c, h)
style_header(ws4, r, len(headers4))

# Top converters
st_sorted = sorted(st_data, key=lambda x: x.get('sales1d', 0), reverse=True)
r = 4
for s in st_sorted[:100]:
    sa = s.get('sales1d', 0)
    co = s['cost']
    acos = (co / sa * 100) if sa > 0 else 0
    roas = (sa / co) if co > 0 else 0

    ws4.cell(r, 1, s['searchTerm'])
    ws4.cell(r, 2, s['campaignName'])
    ws4.cell(r, 3, s.get('targeting', ''))
    ws4.cell(r, 4, s['clicks'])
    ws4.cell(r, 5, round(co)).number_format = MONEY_FMT
    ws4.cell(r, 6, round(sa)).number_format = MONEY_FMT
    ws4.cell(r, 7, s.get('purchases1d', 0))
    ws4.cell(r, 8, round(acos, 1)).number_format = '0.0"%"'
    ws4.cell(r, 9, round(roas, 2)).number_format = '0.00"x"'
    style_row(ws4, r, len(headers4))
    r += 1

# Blank row + wasted terms header
r += 2
ws4.cell(r, 1, "WASTED Search Terms (5+ Clicks, Zero Sales)").font = Font(name="Calibri", bold=True, size=13, color="C00000")
r += 1
for c, h in enumerate(headers4, 1):
    ws4.cell(r, c, h)
style_header(ws4, r, len(headers4))

wasted = [s for s in st_data if s['clicks'] >= 5 and s.get('sales1d', 0) == 0]
wasted.sort(key=lambda x: x['cost'], reverse=True)
r += 1
for s in wasted:
    ws4.cell(r, 1, s['searchTerm'])
    ws4.cell(r, 2, s['campaignName'])
    ws4.cell(r, 3, s.get('targeting', ''))
    ws4.cell(r, 4, s['clicks'])
    ws4.cell(r, 5, round(s['cost'])).number_format = MONEY_FMT
    ws4.cell(r, 6, 0).number_format = MONEY_FMT
    ws4.cell(r, 7, 0)
    ws4.cell(r, 8, "-")
    ws4.cell(r, 9, "-")
    style_row(ws4, r, len(headers4), fill=RED_FILL)
    r += 1

auto_width(ws4)

# ══════════════════════════════════════════════════════════════
# SHEET 5: SD Campaigns
# ══════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("SD Campaigns")
ws5.sheet_properties.tabColor = "C55A11"

headers5 = ["Campaign", "Status", "Spend", "Sales", "Orders", "ACoS", "ROAS", "Clicks", "Impressions", "CTR"]
ws5.cell(1, 1, "Sponsored Display Campaigns — Last 30 Days").font = Font(name="Calibri", bold=True, size=13, color="C55A11")
r = 3
for c, h in enumerate(headers5, 1):
    ws5.cell(r, c, h)
style_header(ws5, r, len(headers5))

sd_sorted = sorted(sd_data, key=lambda x: x.get('sales', 0), reverse=True)
r = 4
for camp in sd_sorted:
    s = camp.get('sales', 0)
    co = camp['cost']
    clicks = camp['clicks']
    imps = camp['impressions']
    orders = camp.get('purchases', 0)
    acos = (co / s * 100) if s > 0 else 0
    roas = (s / co) if co > 0 else 0
    ctr = (clicks / imps * 100) if imps > 0 else 0

    ws5.cell(r, 1, camp['campaignName'])
    ws5.cell(r, 2, camp.get('campaignStatus', ''))
    ws5.cell(r, 3, round(co)).number_format = MONEY_FMT
    ws5.cell(r, 4, round(s)).number_format = MONEY_FMT
    ws5.cell(r, 5, orders)
    ws5.cell(r, 6, round(acos, 1)).number_format = '0.0"%"'
    ws5.cell(r, 7, round(roas, 2)).number_format = '0.00"x"'
    ws5.cell(r, 8, clicks)
    ws5.cell(r, 9, imps)
    ws5.cell(r, 10, round(ctr, 2)).number_format = '0.00"%"'

    if s == 0 and co > 100:
        fill = RED_FILL
    else:
        fill = None
    style_row(ws5, r, len(headers5), fill=fill)
    r += 1

auto_width(ws5)

# ══════════════════════════════════════════════════════════════
# SHEET 6: Issues & Recommendations
# ══════════════════════════════════════════════════════════════
ws6 = wb.create_sheet("Issues & Actions")
ws6.sheet_properties.tabColor = "C00000"

ws6.cell(1, 1, "Issues & Recommended Actions").font = Font(name="Calibri", bold=True, size=14, color="C00000")

# Bleeding campaigns
r = 3
ws6.cell(r, 1, "BLEEDING SP CAMPAIGNS (Spend > ₹200, Zero Sales)").font = BOLD
ws6.cell(r, 1).fill = RED_FILL
r += 1
for c, h in enumerate(["Campaign", "Spend Wasted", "Clicks", "Action"], 1):
    ws6.cell(r, c, h)
style_header(ws6, r, 4)
r += 1
bleeders = [c for c in sp_data if c['cost'] > 200 and c.get('sales1d', 0) == 0]
bleeders.sort(key=lambda x: x['cost'], reverse=True)
for camp in bleeders:
    ws6.cell(r, 1, camp['campaignName'])
    ws6.cell(r, 2, round(camp['cost'])).number_format = MONEY_FMT
    ws6.cell(r, 3, camp['clicks'])
    ws6.cell(r, 4, "PAUSE or fix targeting")
    style_row(ws6, r, 4, fill=RED_FILL)
    r += 1

# High ACoS
r += 2
ws6.cell(r, 1, "HIGH ACoS CAMPAIGNS (ACoS > 30%, Spend > ₹500)").font = BOLD
ws6.cell(r, 1).fill = YELLOW_FILL
r += 1
for c, h in enumerate(["Campaign", "Spend", "Sales", "ACoS", "Action"], 1):
    ws6.cell(r, c, h)
style_header(ws6, r, 5)
r += 1
for camp in sp_data:
    s = camp.get('sales1d', 0)
    co = camp['cost']
    if co > 500 and s > 0:
        ac = co / s * 100
        if ac > 30:
            ws6.cell(r, 1, camp['campaignName'])
            ws6.cell(r, 2, round(co)).number_format = MONEY_FMT
            ws6.cell(r, 3, round(s)).number_format = MONEY_FMT
            ws6.cell(r, 4, round(ac, 1)).number_format = '0.0"%"'
            ws6.cell(r, 5, "Reduce bids / add negatives")
            style_row(ws6, r, 5, fill=YELLOW_FILL)
            r += 1

# Category recommendations
r += 2
ws6.cell(r, 1, "CATEGORY-LEVEL RECOMMENDATIONS").font = Font(name="Calibri", bold=True, size=12, color="2F5496")
r += 1
recs = [
    ("EVA Kids Mat", "✓ ON TARGET — Scale budgets 15-20%. Pause 3 zero-sale ASINs.", GREEN_FILL),
    ("EVA Gym Mat", "✗ Below target (7.6x vs 10x). Clean up negatives. Pause 10 zero-sale ASINs.", RED_FILL),
    ("Storage Bags", "✗ Worst performer (3.3x vs 5.5x). CPC too high at ₹13.84. Restructure or cut budget.", RED_FILL),
    ("Anti-Slip Mats", "~ Close to target (4.1x vs 5x). Optimize bids on underperformers.", YELLOW_FILL),
    ("Baby Play Mat", "✗ Below target (3.9x vs 5.5x). Good organic (57%). Tighten ad keywords.", YELLOW_FILL),
]
for c, h in enumerate(["Category", "Recommendation", "Priority"], 1):
    ws6.cell(r, c, h)
style_header(ws6, r, 3)
r += 1
for cat, rec, fill in recs:
    ws6.cell(r, 1, cat).font = BOLD
    ws6.cell(r, 2, rec)
    ws6.cell(r, 3, "HIGH" if fill == RED_FILL else ("MEDIUM" if fill == YELLOW_FILL else "LOW"))
    style_row(ws6, r, 3, fill=fill)
    r += 1

auto_width(ws6, max_w=80)

# ── Save ──
ts = datetime.now().strftime("%Y%m%d_%H%M")
out_dir = os.path.join(os.path.dirname(__file__), "output")
os.makedirs(out_dir, exist_ok=True)
out_path = os.path.join(out_dir, f"Ads_Performance_Report_{ts}.xlsx")
wb.save(out_path)
print(f"\nExcel report saved to: {out_path}")
print(f"Sheets: {[ws.title for ws in wb.worksheets]}")
