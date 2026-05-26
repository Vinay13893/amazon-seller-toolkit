#!/usr/bin/env python3
"""
Hommies Delivery Timing Excel Generator
========================================
Step 1: Runs amazon_pincode_checker.py for Hommies ASINs → saves raw CSV
Step 2: Converts the raw CSV into a formatted Excel with:
  - Sheet 1: Raw Data (all asin x pincode rows)
  - Sheet 2: Pivot — rows = ASIN, cols = pincode, cell = delivery_type
  - Sheet 3: Summary — count of same_day / next_day / other / unavailable per ASIN

Usage:
  python hommies_delivery_excel.py [--skip-scrape]

  --skip-scrape   Skip the scraping step, just convert existing CSV to Excel.

Output:
  ../output/hommies_delivery_timing.xlsx
"""

import argparse
import subprocess
import sys
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SCRIPT_DIR = Path(__file__).parent
OUTPUT_DIR = SCRIPT_DIR.parent / "output"
RAW_CSV = SCRIPT_DIR / "hommies_availability_report.csv"
OUTPUT_XLSX = OUTPUT_DIR / "hommies_delivery_timing.xlsx"

ASINS_CSV = SCRIPT_DIR / "hommies_asins.csv"
PINCODES_CSV = SCRIPT_DIR / "pincodes.csv"
PROFILE_DIR = SCRIPT_DIR / "amazon_profile"

# Colour map for delivery types
DELIVERY_COLORS = {
    "same_day":   "00B050",  # Green
    "next_day":   "92D050",  # Light green
    "two_day":    "FFEB9C",  # Yellow
    "other":      "FFD966",  # Orange-ish
    "unavailable": "FF0000", # Red
    "":            "FFFFFF",  # White (unknown/error)
}

DELIVERY_LABELS = {
    "same_day":   "Same Day",
    "next_day":   "Next Day",
    "two_day":    "2-Day",
    "other":      "Other",
    "unavailable": "N/A",
    "":            "—",
}


def run_scrape():
    print("Running pincode checker for Hommies ASINs ...")
    cmd = [
        sys.executable,
        str(SCRIPT_DIR / "amazon_pincode_checker.py"),
        "--asins",     str(ASINS_CSV),
        "--pincodes",  str(PINCODES_CSV),
        "--output",    str(RAW_CSV),
        "--profile-dir", str(PROFILE_DIR),
    ]
    result = subprocess.run(cmd)
    if result.returncode != 0:
        print("WARNING: Scraper exited with non-zero code. Proceeding with whatever CSV was written.")


def apply_header_style(ws, row, fill_hex="2F5496"):
    fill = PatternFill("solid", fgColor=fill_hex)
    font = Font(color="FFFFFF", bold=True)
    for cell in ws[row]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def thin_border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)


def build_excel(csv_path: Path, output_path: Path):
    print(f"Reading {csv_path} ...")
    df = pd.read_csv(csv_path)

    # Normalise column names
    df.columns = [c.strip().lower() for c in df.columns]

    # Coerce types
    df["asin"] = df["asin"].astype(str).str.strip().str.upper()
    df["pincode"] = df["pincode"].astype(str).str.strip()
    df["delivery_type"] = df["delivery_type"].fillna("").astype(str).str.strip()

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(str(output_path), engine="openpyxl") as writer:
        # ── Sheet 1: Raw Data ─────────────────────────────────────────────
        raw_cols = ["asin", "pincode", "delivery_type", "delivery_text",
                    "is_buyable", "availability_text", "amazon_fulfilled", "error"]
        raw_cols = [c for c in raw_cols if c in df.columns]
        df[raw_cols].to_excel(writer, sheet_name="Raw Data", index=False)

        # ── Sheet 2: Pivot ────────────────────────────────────────────────
        pivot = df.pivot_table(
            index="asin",
            columns="pincode",
            values="delivery_type",
            aggfunc="first",
            fill_value=""
        )
        pivot.reset_index(inplace=True)
        pivot.to_excel(writer, sheet_name="Delivery by Pincode", index=False)

        # ── Sheet 3: Summary ──────────────────────────────────────────────
        types = ["same_day", "next_day", "two_day", "other", "unavailable"]
        summary_rows = []
        for asin, grp in df.groupby("asin"):
            row = {"ASIN": asin, "Total Pincodes": len(grp)}
            for t in types:
                row[t] = (grp["delivery_type"] == t).sum()
            row["% Fast (Same/Next Day)"] = (
                (row["same_day"] + row["next_day"]) / len(grp) * 100
                if len(grp) else 0
            )
            summary_rows.append(row)
        summary_df = pd.DataFrame(summary_rows).sort_values("% Fast (Same/Next Day)", ascending=False)
        summary_df.to_excel(writer, sheet_name="Summary", index=False)

    # ── Post-processing: colours & column widths ─────────────────────────
    wb = load_workbook(str(output_path))

    # Raw Data sheet
    ws_raw = wb["Raw Data"]
    apply_header_style(ws_raw, 1)
    for col in ws_raw.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws_raw.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 60)
    ws_raw.freeze_panes = "A2"

    # Pivot sheet
    ws_piv = wb["Delivery by Pincode"]
    apply_header_style(ws_piv, 1)
    ws_piv.freeze_panes = "B2"
    # Colour cells by delivery type
    for row in ws_piv.iter_rows(min_row=2):
        for i, cell in enumerate(row):
            if i == 0:  # ASIN column — just bold
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")
                continue
            val = (cell.value or "").lower()
            color = DELIVERY_COLORS.get(val, "FFFFFF")
            label = DELIVERY_LABELS.get(val, str(cell.value))
            cell.fill = PatternFill("solid", fgColor=color)
            cell.value = label
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border()
    # Column widths
    for col in ws_piv.columns:
        ws_piv.column_dimensions[get_column_letter(col[0].column)].width = 12
    ws_piv.column_dimensions["A"].width = 16

    # Summary sheet
    ws_sum = wb["Summary"]
    apply_header_style(ws_sum, 1)
    ws_sum.freeze_panes = "A2"
    for col in ws_sum.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws_sum.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 3, 30)
    # Colour % fast column
    pct_col_idx = None
    for cell in ws_sum[1]:
        if "fast" in str(cell.value or "").lower():
            pct_col_idx = cell.column
            break
    if pct_col_idx:
        for row in ws_sum.iter_rows(min_row=2, min_col=pct_col_idx, max_col=pct_col_idx):
            for cell in row:
                try:
                    pct = float(cell.value or 0)
                    if pct >= 50:
                        cell.fill = PatternFill("solid", fgColor="00B050")
                    elif pct >= 25:
                        cell.fill = PatternFill("solid", fgColor="FFEB9C")
                    else:
                        cell.fill = PatternFill("solid", fgColor="FF0000")
                        cell.font = Font(color="FFFFFF")
                    cell.number_format = "0.0%"
                    cell.value = pct / 100
                except (TypeError, ValueError):
                    pass

    wb.save(str(output_path))
    print(f"\n✓ Excel saved → {output_path}")
    print(f"  Sheets: Raw Data | Delivery by Pincode (pivot) | Summary")


def main():
    parser = argparse.ArgumentParser(description="Hommies Delivery Timing Excel")
    parser.add_argument("--skip-scrape", action="store_true",
                        help="Skip scraping, just convert existing CSV to Excel")
    args = parser.parse_args()

    if not args.skip_scrape:
        run_scrape()
    else:
        print("Skipping scrape — using existing CSV.")

    if not RAW_CSV.exists():
        print(f"ERROR: {RAW_CSV} not found. Run without --skip-scrape first.")
        sys.exit(1)

    build_excel(RAW_CSV, OUTPUT_XLSX)


if __name__ == "__main__":
    main()
