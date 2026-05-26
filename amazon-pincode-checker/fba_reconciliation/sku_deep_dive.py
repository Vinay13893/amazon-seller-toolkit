"""
SKU Deep-Dive — Damaged Returns & Reimbursement Recovery Analysis
=================================================================
Usage:
    python sku_deep_dive.py insulation_cover_1000
    python sku_deep_dive.py insulation_cover_1000 --days 90

Pulls:
  1. All returns for this SKU (full disposition + reason + order)
  2. All reimbursements for this SKU
  3. Inventory Adjustments (lost / damaged / found events)
  4. Full Inventory Ledger (all stock movements)
  5. Cross-references each CARRIER_DAMAGED / WAREHOUSE_DAMAGED return
     against reimbursements to find gaps
"""

import argparse
import os
import sys
import time
from datetime import datetime, timezone, timedelta

import pandas as pd

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT_DIR = os.path.join(SCRIPT_DIR, "output")
os.makedirs(OUTPUT_DIR, exist_ok=True)

sys.path.insert(0, SCRIPT_DIR)
from sp_client import (
    _run_report, _call,
    MARKETPLACE_ID,
)


# ─── Adjustments Report ───────────────────────────────────────
def fetch_adjustments(start_dt, end_dt):
    """GET_FBA_FULFILLMENT_INVENTORY_ADJUSTMENTS_DATA — every inventory
    adjustment: LOST, FOUND, DAMAGED, MISMATCH, DISPOSED, etc."""
    print(f"\n  [3] Inventory Adjustments ({start_dt.date()} → {end_dt.date()}) ...")
    raw = _run_report(
        "GET_FBA_FULFILLMENT_INVENTORY_ADJUSTMENTS_DATA", start_dt, end_dt
    )
    print(f"    Rows: {len(raw)}, Columns: {list(raw.columns)}")
    return raw


def fetch_ledger(start_dt, end_dt):
    """GET_LEDGER_DETAIL_VIEW_DATA — full per-FNSKU event stream."""
    print(f"\n  [4] Inventory Ledger ({start_dt.date()} → {end_dt.date()}) ...")
    chunks = []
    cur = start_dt
    while cur <= end_dt:
        nxt = min(cur + timedelta(days=30), end_dt)
        print(f"    Chunk: {cur.date()} → {nxt.date()}")
        df = _run_report("GET_LEDGER_DETAIL_VIEW_DATA", cur, nxt)
        chunks.append(df)
        cur = nxt + timedelta(days=1)
    raw = pd.concat(chunks, ignore_index=True) if len(chunks) > 1 else chunks[0]
    print(f"    Total rows: {len(raw)}")
    return raw


def fetch_returns(start_dt, end_dt):
    print(f"\n  [1] Customer Returns ({start_dt.date()} → {end_dt.date()}) ...")
    raw = _run_report(
        "GET_FBA_FULFILLMENT_CUSTOMER_RETURNS_DATA", start_dt, end_dt
    )
    print(f"    Rows: {len(raw)}")
    return raw


def fetch_reimbursements(start_dt, end_dt):
    print(f"\n  [2] Reimbursements ({start_dt.date()} → {end_dt.date()}) ...")
    raw = _run_report("GET_FBA_REIMBURSEMENTS_DATA", start_dt, end_dt)
    print(f"    Rows: {len(raw)}")
    return raw


# ─── Analysis ─────────────────────────────────────────────────
def _col(df, *cands):
    for c in cands:
        if c in df.columns:
            return c
    return None


def analyse_sku(sku, returns_raw, reimb_raw, adj_raw, ledger_raw):
    sku_lower = sku.lower()

    # ── Filter to this SKU ────────────────────────────────────
    def _filter(df, *sku_cols):
        if df is None or df.empty:
            return pd.DataFrame()
        for col in sku_cols:
            if col in df.columns:
                mask = df[col].astype(str).str.lower() == sku_lower
                if mask.any():
                    return df[mask].copy()
        # fallback: search all columns
        mask = df.apply(
            lambda c: c.astype(str).str.lower() == sku_lower
        ).any(axis=1)
        return df[mask].copy()

    ret  = _filter(returns_raw,  "sku", "seller-sku", "msku")
    rei  = _filter(reimb_raw,    "sku", "seller-sku", "msku")
    adj  = _filter(adj_raw,      "sku", "seller-sku", "msku") if adj_raw is not None and not adj_raw.empty else pd.DataFrame()
    led  = _filter(ledger_raw,   "msku", "sku", "seller-sku") if ledger_raw is not None and not ledger_raw.empty else pd.DataFrame()

    print(f"\n  SKU '{sku}': returns={len(ret)}, reimbursements={len(rei)}, adjustments={len(adj)}, ledger={len(led)}")

    # ─── 1. Returns breakdown ─────────────────────────────────
    disp_col   = _col(ret, "detailed-disposition", "disposition")
    reason_col = _col(ret, "reason")
    order_col  = _col(ret, "order-id", "amazon-order-id")
    date_col   = _col(ret, "return-date", "date")
    fc_col     = _col(ret, "fulfillment-center-id", "fulfillment-center")

    if not ret.empty and disp_col:
        ret_clean = pd.DataFrame({
            "return_date":  pd.to_datetime(ret[date_col],   errors="coerce") if date_col   else None,
            "order_id":     ret[order_col].str.strip()                        if order_col  else "",
            "disposition":  ret[disp_col].str.strip().str.upper(),
            "reason":       ret[reason_col].str.strip()                       if reason_col else "",
            "fc":           ret[fc_col].str.strip()                           if fc_col     else "",
            "qty": 1,
        })
    else:
        ret_clean = pd.DataFrame()

    # ─── 2. Reimbursements ────────────────────────────────────
    reimb_date  = _col(rei, "approval-date", "date")
    reimb_amt   = _col(rei, "amount-total",  "amount-per-unit", "amount")
    reimb_rsn   = _col(rei, "reason")
    reimb_order = _col(rei, "amazon-order-id", "order-id")
    reimb_qcash = _col(rei, "quantity-reimbursed-cash")
    reimb_qinv  = _col(rei, "quantity-reimbursed-inventory")
    reimb_qtot  = _col(rei, "quantity-reimbursed-total")

    if not rei.empty:
        rei_clean = pd.DataFrame({
            "approval_date": pd.to_datetime(rei[reimb_date],  errors="coerce") if reimb_date  else None,
            "reason":        rei[reimb_rsn].str.strip()                         if reimb_rsn  else "",
            "order_id":      rei[reimb_order].str.strip()                       if reimb_order else "",
            "amount":        pd.to_numeric(rei[reimb_amt],   errors="coerce").fillna(0) if reimb_amt  else 0,
            "qty_cash":      pd.to_numeric(rei[reimb_qcash], errors="coerce").fillna(0).astype(int) if reimb_qcash else 0,
            "qty_inventory": pd.to_numeric(rei[reimb_qinv],  errors="coerce").fillna(0).astype(int) if reimb_qinv  else 0,
            "qty_total":     pd.to_numeric(rei[reimb_qtot],  errors="coerce").fillna(0).astype(int) if reimb_qtot  else 0,
        })
    else:
        rei_clean = pd.DataFrame()

    # ─── 3. Adjustments ──────────────────────────────────────
    adj_clean = pd.DataFrame()
    if not adj.empty:
        adj_date = _col(adj, "date", "adjusted-date", "snapshot-date")
        adj_rsn  = _col(adj, "reason", "adjustment-reason")
        adj_qty  = _col(adj, "quantity", "qty")
        adj_disp = _col(adj, "disposition")
        adj_fnsku= _col(adj, "fnsku")

        adj_clean = pd.DataFrame({
            "date":        pd.to_datetime(adj[adj_date], errors="coerce") if adj_date else None,
            "reason":      adj[adj_rsn].str.strip()                        if adj_rsn  else "",
            "disposition": adj[adj_disp].str.strip()                       if adj_disp else "",
            "fnsku":       adj[adj_fnsku].str.strip()                      if adj_fnsku else "",
            "qty":         pd.to_numeric(adj[adj_qty], errors="coerce").fillna(0).astype(int) if adj_qty else 0,
        })

    # ─── 4. Ledger events ─────────────────────────────────────
    led_clean = pd.DataFrame()
    if not led.empty:
        led_date  = _col(led, "date", "snapshot-date")
        led_evt   = _col(led, "event-type", "transaction-type")
        led_qty   = _col(led, "quantity", "qty")
        led_disp  = _col(led, "disposition")
        led_ref   = _col(led, "reference-id", "shipment-id")

        led_clean = pd.DataFrame({
            "date":       pd.to_datetime(led[led_date], errors="coerce") if led_date else None,
            "event_type": led[led_evt].str.strip()                        if led_evt  else "",
            "disposition":led[led_disp].str.strip()                       if led_disp else "",
            "reference":  led[led_ref].str.strip()                        if led_ref  else "",
            "qty":        pd.to_numeric(led[led_qty], errors="coerce").fillna(0).astype(int) if led_qty else 0,
        })

    # ─── 5. Cross-reference: damaged returns vs reimbursements ─
    cross = pd.DataFrame()
    if not ret_clean.empty and not rei_clean.empty:
        # Amazon should reimburse for CARRIER_DAMAGED and WAREHOUSE_DAMAGED
        # Match by order_id where possible
        claimable = ret_clean[
            ret_clean["disposition"].isin(["CARRIER_DAMAGED", "WAREHOUSE_DAMAGED"])
        ].copy()

        # Reimbursements with reason CustomerReturn or LostInboundShipment
        relevant_reimb = rei_clean.copy()
        reimb_orders   = set(relevant_reimb["order_id"].dropna().unique())
        claimable["reimbursed"] = claimable["order_id"].isin(reimb_orders)
        if rei_clean.empty:
            claimable["reimb_amount"] = 0.0
        else:
            order_amt = rei_clean.groupby("order_id")["amount"].sum().to_dict()
            claimable["reimb_amount"] = claimable["order_id"].map(order_amt).fillna(0)

        claimable["claim_status"] = claimable.apply(
            lambda r: (
                "✓ Reimbursed" if r["reimbursed"] and r["reimb_amount"] > 0
                else "⚠ REIMBURSED ₹0 — CHECK!" if r["reimbursed"]
                else "✗ NOT REIMBURSED — FILE CLAIM"
            ), axis=1
        )
        cross = claimable

    # ─── 6. Summary numbers ──────────────────────────────────
    total_returns      = len(ret_clean)                                             if not ret_clean.empty else 0
    sellable_returns   = (ret_clean["disposition"] == "SELLABLE").sum()             if not ret_clean.empty else 0
    carrier_damaged    = (ret_clean["disposition"] == "CARRIER_DAMAGED").sum()      if not ret_clean.empty else 0
    warehouse_damaged  = (ret_clean["disposition"] == "WAREHOUSE_DAMAGED").sum()    if not ret_clean.empty else 0
    customer_damaged   = (ret_clean["disposition"] == "CUSTOMER_DAMAGED").sum()     if not ret_clean.empty else 0
    total_reimb_amount = rei_clean["amount"].sum()                                  if not rei_clean.empty else 0
    total_reimb_units  = rei_clean["qty_total"].sum()                               if not rei_clean.empty else 0
    not_reimbursed     = (cross["claim_status"] == "✗ NOT REIMBURSED — FILE CLAIM").sum() if not cross.empty else 0

    adj_lost    = 0
    adj_damaged = 0
    adj_found   = 0
    if not adj_clean.empty:
        rsn = adj_clean["reason"].str.upper()
        adj_lost    = adj_clean.loc[rsn.str.contains("LOST|MISSING",   na=False), "qty"].sum()
        adj_damaged = adj_clean.loc[rsn.str.contains("DAMAG|DEFECT",   na=False), "qty"].sum()
        adj_found   = adj_clean.loc[rsn.str.contains("FOUND|RECOVER",  na=False), "qty"].sum()

    summary = {
        "sku":                     sku,
        "total_returns":           total_returns,
        "sellable_returns":        sellable_returns,
        "carrier_damaged":         carrier_damaged,
        "warehouse_damaged":       warehouse_damaged,
        "customer_damaged":        customer_damaged,
        "total_unsellable":        total_returns - sellable_returns,
        "total_reimb_amount":      round(total_reimb_amount, 2),
        "total_reimb_units":       int(total_reimb_units),
        "per_unit_reimb":          round(total_reimb_amount / total_reimb_units, 2) if total_reimb_units > 0 else 0,
        "not_reimbursed_claims":   int(not_reimbursed),
        "adj_lost_units":          int(adj_lost),
        "adj_damaged_units":       int(adj_damaged),
        "adj_found_units":         int(adj_found),
    }

    print(f"\n{'='*60}")
    print(f"  DEEP-DIVE SUMMARY: {sku}")
    print(f"{'='*60}")
    print(f"  Total Returns:           {summary['total_returns']}")
    print(f"    Sellable (back):       {summary['sellable_returns']}")
    print(f"    Carrier Damaged:       {summary['carrier_damaged']}  ← Amazon carrier fault, should be reimbursed")
    print(f"    Warehouse Damaged:     {summary['warehouse_damaged']} ← Amazon FC fault, should be reimbursed")
    print(f"    Customer Damaged:      {summary['customer_damaged']} ← Customer fault")
    print(f"  Total Reimbursed:        ₹{summary['total_reimb_amount']:,.2f}  ({summary['total_reimb_units']} units)")
    print(f"  Per-Unit Reimbursement:  ₹{summary['per_unit_reimb']:,.2f}")
    print(f"  Unreimbursed Claims:     {summary['not_reimbursed_claims']} units  ← FILE CLAIMS")
    if not adj_clean.empty:
        print(f"\n  Inventory Adjustments:")
        print(f"    Lost/Missing:          {adj_lost}")
        print(f"    Damaged:               {adj_damaged}")
        print(f"    Found/Recovered:       {adj_found}")
    print(f"{'='*60}")

    return {
        "summary":      pd.DataFrame([summary]),
        "returns":      ret_clean,
        "reimbursements": rei_clean,
        "adjustments":  adj_clean,
        "ledger":       led_clean,
        "cross_check":  cross,
    }


# ─── Excel Output ─────────────────────────────────────────────
def write_excel(sku, results, output_path):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    CLR_HEADER = "1F4E79"
    CLR_OK     = "C6EFCE"
    CLR_WARN   = "FFEB9C"
    CLR_ALERT  = "FFC7CE"
    CLR_BLUE   = "D6E4F0"

    THIN   = Side(style="thin", color="B0B0B0")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    def _strip_tz(df):
        df = df.copy()
        for col in df.columns:
            if pd.api.types.is_datetime64_any_dtype(df[col]):
                try:    df[col] = df[col].dt.tz_localize(None)
                except: df[col] = df[col].dt.tz_convert(None)
        return df

    writer = pd.ExcelWriter(output_path, engine="openpyxl")

    sheets = [
        ("🔍 Cross-Check",    results["cross_check"]),
        ("📋 Returns",         results["returns"]),
        ("💰 Reimbursements",  results["reimbursements"]),
        ("🔧 Adjustments",     results["adjustments"]),
        ("📒 Ledger",          results["ledger"]),
        ("📊 Summary",         results["summary"]),
    ]

    for sheet_name, df in sheets:
        if df is None or df.empty:
            df = pd.DataFrame({"note": ["No data for this section"]})
        df = _strip_tz(df)
        df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=1)
        ws = writer.sheets[sheet_name]

        # Title banner
        title = f"SKU Deep-Dive: {sku}   |   {sheet_name}"
        ws.cell(row=1, column=1, value=title).font = Font(bold=True, size=12, color=CLR_HEADER)
        ws.cell(row=1, column=1).fill = PatternFill("solid", fgColor=CLR_BLUE)
        if len(df.columns) > 1:
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))

        # Header row
        for ci in range(1, len(df.columns) + 1):
            c = ws.cell(row=2, column=ci)
            c.font      = Font(bold=True, color="FFFFFF", size=10)
            c.fill      = PatternFill("solid", fgColor=CLR_HEADER)
            c.alignment = Alignment(horizontal="center", wrap_text=True)
            c.border    = BORDER

        # Data rows — colour by claim_status or disposition
        for ri in range(3, ws.max_row + 1):
            for ci in range(1, len(df.columns) + 1):
                c = ws.cell(row=ri, column=ci)
                c.font      = Font(size=10)
                c.alignment = Alignment(horizontal="left")
                c.border    = BORDER

            # Colour cross-check sheet by claim status
            if sheet_name == "🔍 Cross-Check" and "claim_status" in df.columns:
                ci_status = list(df.columns).index("claim_status") + 1
                val = ws.cell(row=ri, column=ci_status).value or ""
                color = CLR_OK if "Reimbursed" in str(val) else CLR_ALERT
                for ci in range(1, len(df.columns) + 1):
                    ws.cell(row=ri, column=ci).fill = PatternFill("solid", fgColor=color)

            # Colour returns sheet by disposition
            if sheet_name == "📋 Returns" and "disposition" in df.columns:
                ci_d = list(df.columns).index("disposition") + 1
                val  = str(ws.cell(row=ri, column=ci_d).value or "")
                color = (CLR_OK    if val == "SELLABLE"
                         else CLR_ALERT if val in ("CARRIER_DAMAGED", "WAREHOUSE_DAMAGED")
                         else CLR_WARN)
                for ci in range(1, len(df.columns) + 1):
                    ws.cell(row=ri, column=ci).fill = PatternFill("solid", fgColor=color)

        # Auto-width
        for col_cells in ws.columns:
            max_len    = max((len(str(c.value or "")) for c in col_cells), default=8)
            col_letter = get_column_letter(col_cells[0].column)
            ws.column_dimensions[col_letter].width = max(12, min(max_len + 2, 55))

        ws.freeze_panes = ws.cell(row=3, column=1)

    writer.close()
    print(f"\n  ✓ Saved: {output_path}")


# ─── Main ─────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description="SKU Deep-Dive: damaged returns & reimbursement recovery")
    parser.add_argument("sku",        help="SKU to analyse e.g. insulation_cover_1000")
    parser.add_argument("--days",     type=int, default=90, help="Lookback days (default 90)")
    parser.add_argument("--no-ledger",action="store_true",  help="Skip full inventory ledger (faster)")
    args = parser.parse_args()

    end_dt   = datetime.now(timezone.utc).replace(hour=0, minute=0, second=0, microsecond=0) - timedelta(seconds=1)
    start_dt = (end_dt - timedelta(days=args.days - 1)).replace(hour=0, minute=0, second=0, microsecond=0)

    print(f"\n{'='*60}")
    print(f"  SKU Deep-Dive: {args.sku}")
    print(f"  Period: {start_dt.date()} → {end_dt.date()} ({args.days} days)")
    print(f"{'='*60}")

    returns_raw     = pd.DataFrame()
    reimb_raw       = pd.DataFrame()
    adj_raw         = pd.DataFrame()
    ledger_raw      = pd.DataFrame()

    try:
        returns_raw = fetch_returns(start_dt, end_dt)
    except Exception as e:
        print(f"  ⚠ Returns failed: {e}")

    try:
        reimb_raw = fetch_reimbursements(start_dt, end_dt)
    except Exception as e:
        print(f"  ⚠ Reimbursements failed: {e}")

    try:
        adj_raw = fetch_adjustments(start_dt, end_dt)
    except Exception as e:
        print(f"  ⚠ Adjustments failed: {e}")

    if not args.no_ledger:
        try:
            ledger_raw = fetch_ledger(start_dt, end_dt)
        except Exception as e:
            print(f"  ⚠ Ledger failed: {e}")
    else:
        print("\n  [4] Ledger skipped (--no-ledger)")

    results = analyse_sku(args.sku, returns_raw, reimb_raw, adj_raw, ledger_raw)

    ts          = datetime.now().strftime("%Y%m%d_%H%M")
    output_path = os.path.join(OUTPUT_DIR, f"deep_dive_{args.sku}_{args.days}d_{ts}.xlsx")
    write_excel(args.sku, results, output_path)
    print(f"\n  Done.")


if __name__ == "__main__":
    main()
