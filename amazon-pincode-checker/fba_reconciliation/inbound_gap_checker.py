"""
FBA Inbound Gap Checker
=======================
Checks whether all stock you sent to Amazon was fully received at the FC.

Data sources:
  1. Inbound Shipments API  — QuantityShipped vs QuantityReceived per SKU per shipment
     (uses QueryType=DATE_RANGE so no ShipmentStatusList required — fixes the 400 bug)
  2. Inventory Ledger       — every "Receipts" event confirms actual FC inward
  3. Inventory Snapshot     — current "inbound" qty (in-transit / being received)

Usage:
    python inbound_gap_checker.py
    python inbound_gap_checker.py --days 90
    python inbound_gap_checker.py --days 180
"""

import argparse
import gzip
import hashlib
import hmac
import json
import os
import sys
import time
from datetime import datetime, timedelta, timezone
from io import StringIO

import requests as req
import pandas as pd

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT_DIR = os.path.join(SCRIPT_DIR, "output")
os.makedirs(OUTPUT_DIR, exist_ok=True)

sys.path.insert(0, SCRIPT_DIR)
from sp_client import (
    _load_config, _token, _call, _run_report,
    MARKETPLACE_ID, SP_API_HOST, SP_API_BASE, SP_API_REGION,
    _signature_key,
    fetch_inventory_snapshot,
)


# ─── Inbound Shipments via DATE_RANGE (fixes the 400 bug) ─────
def fetch_inbound_shipments_date_range(last_updated_after: datetime):
    """
    Uses QueryType=DATE_RANGE + LastUpdatedAfter — no ShipmentStatusList required.
    Returns list of shipment dicts.
    """
    shipments  = []
    next_token = None
    page       = 0

    print(f"\n  [1] Inbound Shipments (updated after {last_updated_after.date()}) ...")

    while True:
        page += 1
        if next_token:
            params = {
                "MarketplaceId": MARKETPLACE_ID,
                "QueryType":     "NEXT_TOKEN",
                "NextToken":     next_token,
            }
        else:
            now_str = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
            params = {
                "MarketplaceId":     MARKETPLACE_ID,
                "QueryType":         "DATE_RANGE",
                "LastUpdatedAfter":  last_updated_after.strftime("%Y-%m-%dT00:00:00Z"),
                "LastUpdatedBefore": now_str,
            }

        try:
            res = _call("GET", "/fba/inbound/v0/shipments", params=params)
        except RuntimeError as e:
            print(f"    ⚠ Inbound API error: {e}")
            return []

        payload    = res.get("payload", {})
        items      = payload.get("ShipmentData", [])
        next_token = payload.get("NextToken")

        for s in items:
            shipments.append({
                "shipment_id":    s.get("ShipmentId", ""),
                "shipment_name":  s.get("ShipmentName", ""),
                "status":         s.get("ShipmentStatus", ""),
                "destination_fc": s.get("DestinationFulfillmentCenterId", ""),
                "label_type":     s.get("LabelPrepType", ""),
            })

        print(f"    Page {page}: {len(items)} shipments")
        if not next_token or not items:
            break
        time.sleep(1)

    print(f"    Total shipments found: {len(shipments)}")
    return shipments


def fetch_shipment_items(shipments):
    """
    For each shipment, call /fba/inbound/v0/shipments/{id}/items to get
    QuantityShipped and QuantityReceived per SKU.
    """
    rows = []
    print(f"\n  [2] Fetching items for {len(shipments)} shipments ...")
    for i, s in enumerate(shipments):
        sid = s["shipment_id"]
        try:
            time.sleep(0.4)
            res = _call("GET", f"/fba/inbound/v0/shipments/{sid}/items",
                        params={"MarketplaceId": MARKETPLACE_ID})
            items = res.get("payload", {}).get("ItemData", [])
            for item in items:
                rows.append({
                    "shipment_id":    sid,
                    "shipment_name":  s["shipment_name"],
                    "status":         s["status"],
                    "destination_fc": s["destination_fc"],
                    "sku":            item.get("SellerSKU", ""),
                    "fnsku":          item.get("FulfillmentNetworkSKU", ""),
                    "qty_shipped":    int(item.get("QuantityShipped",  0) or 0),
                    "qty_received":   int(item.get("QuantityReceived", 0) or 0),
                })
        except Exception as e:
            print(f"    ⚠ Could not fetch items for {sid}: {e}")
        if (i + 1) % 10 == 0:
            print(f"    ... {i+1}/{len(shipments)} done")

    df = pd.DataFrame(rows)
    if not df.empty:
        df["gap"] = df["qty_received"] - df["qty_shipped"]
    print(f"    Total line items: {len(df)}")
    return df


# ─── Inventory Ledger Receipts ─────────────────────────────────
def fetch_ledger_receipts(start_dt, end_dt):
    """
    Pull the full inventory ledger, filter to 'Receipts' event type.
    These are confirmed inwards — units Amazon actually put into stock.
    """
    print(f"\n  [3] Inventory Ledger Receipts ({start_dt.date()} → {end_dt.date()}) ...")
    chunks    = []
    cur_start = start_dt
    while cur_start <= end_dt:
        cur_end = min(cur_start + timedelta(days=30), end_dt)
        print(f"    Chunk: {cur_start.date()} → {cur_end.date()}")
        df = _run_report("GET_LEDGER_DETAIL_VIEW_DATA", cur_start, cur_end)
        chunks.append(df)
        cur_start = cur_end + timedelta(days=1)

    raw = pd.concat(chunks, ignore_index=True) if len(chunks) > 1 else chunks[0]
    print(f"    Total ledger rows: {len(raw)}")

    # Find column names
    def col(df, *c):
        for x in c:
            if x in df.columns:
                return x
        return None

    evt_col  = col(raw, "event-type", "event_type", "transaction-type")
    sku_col  = col(raw, "msku", "sku", "seller-sku")
    qty_col  = col(raw, "quantity", "qty")
    date_col = col(raw, "date", "snapshot-date")
    ref_col  = col(raw, "reference-id", "reference_id", "shipment-id")
    fc_col   = col(raw, "fulfillment-center", "fulfillment-center-id")
    disp_col = col(raw, "disposition")

    print(f"    Ledger columns: {list(raw.columns)}")

    # Filter to Receipts only
    if evt_col:
        receipts = raw[raw[evt_col].str.upper().str.contains("RECEIPT", na=False)].copy()
    else:
        receipts = raw.copy()  # can't filter, show all

    print(f"    Receipt rows: {len(receipts)}")

    out = pd.DataFrame({
        "date":         pd.to_datetime(receipts[date_col], errors="coerce") if date_col else pd.NaT,
        "sku":          receipts[sku_col].str.strip()                        if sku_col  else "",
        "event_type":   receipts[evt_col].str.strip()                        if evt_col  else "",
        "disposition":  receipts[disp_col].str.strip()                       if disp_col else "",
        "fc":           receipts[fc_col].str.strip()                          if fc_col   else "",
        "reference_id": receipts[ref_col].str.strip()                         if ref_col  else "",
        "qty_received": pd.to_numeric(receipts[qty_col], errors="coerce").fillna(0).astype(int) if qty_col else 0,
    })
    return out, raw


# ─── Analysis ─────────────────────────────────────────────────
def analyse_gaps(shipment_items, ledger_receipts, inventory_snapshot):
    """
    Cross-reference shipped vs received at three levels:
      A) Shipment level: sum(qty_shipped) vs sum(qty_received) per shipment
      B) SKU level:      sum(qty_shipped) vs sum(qty_received) per SKU across all shipments
      C) Ledger check:   qty_received from ledger vs from shipment items (should match)
      D) Current inbound: SKUs with inbound > 0 in snapshot (still awaiting receipt)
    """
    results = {}

    # ── A: Shipment-level summary ─────────────────────────────
    if not shipment_items.empty:
        ship_summary = (
            shipment_items
            .groupby(["shipment_id", "shipment_name", "status", "destination_fc"])
            .agg(
                total_shipped  = ("qty_shipped",  "sum"),
                total_received = ("qty_received", "sum"),
                sku_count      = ("sku",          "nunique"),
            )
            .reset_index()
        )
        ship_summary["total_gap"]      = ship_summary["total_received"] - ship_summary["total_shipped"]
        ship_summary["pct_received"]   = (
            (ship_summary["total_received"] / ship_summary["total_shipped"].replace(0, 1)) * 100
        ).round(1)
        ship_summary["gap_flag"] = ship_summary.apply(
            lambda r: (
                "✓ Complete"          if r["total_gap"] == 0
                else "⚠ SHORT INWARD" if r["total_gap"] < 0
                else "⚠ OVER RECEIVED" if r["total_gap"] > 0
                else "⏳ Pending"
            ), axis=1
        )
        # Sort: short inward first, then by shipment name
        ship_summary = ship_summary.sort_values(
            ["gap_flag", "shipment_name"], ascending=[False, True]
        ).reset_index(drop=True)
        results["shipment_summary"] = ship_summary

        short_inward = ship_summary[ship_summary["total_gap"] < 0]
        print(f"\n  Shipment summary:")
        print(f"    Total shipments:           {len(ship_summary)}")
        print(f"    Fully received:            {(ship_summary['total_gap'] == 0).sum()}")
        print(f"    Short inward (gaps):       {len(short_inward)} ← investigate these")
        print(f"    Over received (unusual):   {(ship_summary['total_gap'] > 0).sum()}")

        if not short_inward.empty:
            print("\n  SHORT INWARD SHIPMENTS:")
            for _, r in short_inward.iterrows():
                gap_units = abs(r["total_gap"])
                print(f"    {r['shipment_id']} | {r['status']:10} | FC: {r['destination_fc']:6} "
                      f"| Sent: {r['total_shipped']:4} | Rcvd: {r['total_received']:4} "
                      f"| GAP: -{gap_units} units  ← FILE CLAIM")
    else:
        print("\n  ⚠ No shipment item data — inbound API unavailable")
        results["shipment_summary"] = pd.DataFrame()

    # ── B: SKU-level gaps from shipment data ─────────────────
    if not shipment_items.empty:
        sku_summary = (
            shipment_items
            .groupby("sku")
            .agg(
                total_shipped  = ("qty_shipped",  "sum"),
                total_received = ("qty_received", "sum"),
                shipment_count = ("shipment_id",  "nunique"),
            )
            .reset_index()
        )
        sku_summary["gap"]       = sku_summary["total_received"] - sku_summary["total_shipped"]
        sku_summary["gap_flag"]  = sku_summary["gap"].apply(
            lambda g: "✓ OK" if g == 0 else ("⚠ SHORT" if g < 0 else "⚠ OVER")
        )
        sku_summary = sku_summary.sort_values("gap").reset_index(drop=True)
        results["sku_gaps"] = sku_summary

        sku_gaps = sku_summary[sku_summary["gap"] < 0]
        if not sku_gaps.empty:
            print(f"\n  SKUs with inward shortages ({len(sku_gaps)} SKUs):")
            for _, r in sku_gaps.iterrows():
                print(f"    {r['sku']:40s} | Sent: {r['total_shipped']:5} | Rcvd: {r['total_received']:5} | Gap: {r['gap']:5}")
        else:
            print("\n  No SKU-level inward shortages found.")
    else:
        results["sku_gaps"] = pd.DataFrame()

    # ── C: Ledger Receipts cross-check ───────────────────────
    if not ledger_receipts.empty:
        ledger_by_sku = (
            ledger_receipts
            .groupby("sku")
            .agg(
                ledger_received = ("qty_received", "sum"),
                receipt_events  = ("qty_received", "count"),
                fcs             = ("fc",           lambda x: ", ".join(sorted(x.dropna().unique()))),
            )
            .reset_index()
            .sort_values("ledger_received", ascending=False)
        )
        results["ledger_receipts"] = ledger_receipts
        results["ledger_by_sku"]   = ledger_by_sku
        print(f"\n  Ledger Receipts: {len(ledger_by_sku)} unique SKUs received stock")
        print(f"    Total units inwarded (ledger): {ledger_receipts['qty_received'].sum()}")

        # Cross-reference with shipment data if available
        if not shipment_items.empty and "sku_gaps" in results and not results["sku_gaps"].empty:
            merged = results["sku_gaps"].merge(
                ledger_by_sku[["sku", "ledger_received"]], on="sku", how="outer"
            ).fillna(0)
            merged["ledger_gap"] = merged["ledger_received"] - merged["total_shipped"]
            results["cross_check"] = merged
    else:
        results["ledger_receipts"] = pd.DataFrame()
        results["ledger_by_sku"]   = pd.DataFrame()

    # ── D: Current inbound from snapshot ─────────────────────
    if not inventory_snapshot.empty:
        currently_inbound = inventory_snapshot[inventory_snapshot["inbound"] > 0].copy()
        currently_inbound = currently_inbound.sort_values("inbound", ascending=False).reset_index(drop=True)
        results["currently_inbound"] = currently_inbound
        print(f"\n  Currently inbound (in transit / being received): {len(currently_inbound)} SKUs, "
              f"{currently_inbound['inbound'].sum()} total units")
    else:
        results["currently_inbound"] = pd.DataFrame()

    return results


# ─── Excel Output ──────────────────────────────────────────────
def write_excel(results, output_path, period_label):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    CLR_HDR   = "1F4E79"
    CLR_OK    = "C6EFCE"
    CLR_WARN  = "FFEB9C"
    CLR_ALERT = "FFC7CE"
    CLR_BLUE  = "D6E4F0"
    THIN      = Side(style="thin", color="B0B0B0")
    BRD       = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    def _strip_tz(df):
        df = df.copy()
        for col in df.columns:
            if pd.api.types.is_datetime64_any_dtype(df[col]):
                try:    df[col] = df[col].dt.tz_localize(None)
                except: df[col] = df[col].dt.tz_convert(None)
        return df

    def _write_sheet(writer, name, df, flag_col=None, ok_val="✓"):
        if df is None or df.empty:
            df = pd.DataFrame({"note": [f"No data — {name}"]})
        df = _strip_tz(df)
        df.to_excel(writer, sheet_name=name, index=False, startrow=1)
        ws = writer.sheets[name]

        # Title
        title = f"FBA Inbound Gap Checker | {period_label}   |   {name}"
        ws.cell(row=1, column=1, value=title).font  = Font(bold=True, size=12, color=CLR_HDR)
        ws.cell(row=1, column=1).fill = PatternFill("solid", fgColor=CLR_BLUE)
        if len(df.columns) > 1:
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))

        # Header
        for ci in range(1, len(df.columns) + 1):
            c = ws.cell(row=2, column=ci)
            c.font      = Font(bold=True, color="FFFFFF", size=10)
            c.fill      = PatternFill("solid", fgColor=CLR_HDR)
            c.alignment = Alignment(horizontal="center", wrap_text=True)
            c.border    = BRD

        # Data rows
        for ri in range(3, ws.max_row + 1):
            row_color = None
            if flag_col and flag_col in df.columns:
                ci_flag = list(df.columns).index(flag_col) + 1
                val     = str(ws.cell(row=ri, column=ci_flag).value or "")
                row_color = (
                    CLR_OK    if ok_val in val
                    else CLR_ALERT if "SHORT" in val or "GAP" in val or "OVER" in val
                    else CLR_WARN
                )
            for ci in range(1, len(df.columns) + 1):
                c           = ws.cell(row=ri, column=ci)
                c.font      = Font(size=10)
                c.alignment = Alignment(horizontal="left")
                c.border    = BRD
                if row_color:
                    c.fill = PatternFill("solid", fgColor=row_color)

        # Auto-width
        for col_cells in ws.columns:
            w = max((len(str(c.value or "")) for c in col_cells), default=8)
            ws.column_dimensions[get_column_letter(col_cells[0].column)].width = max(12, min(w + 2, 60))
        ws.freeze_panes = ws.cell(row=3, column=1)

    writer = pd.ExcelWriter(output_path, engine="openpyxl")

    _write_sheet(writer, "🚚 Shipment Summary",  results.get("shipment_summary", pd.DataFrame()), "gap_flag", "✓")
    _write_sheet(writer, "📦 SKU Gaps",           results.get("sku_gaps",        pd.DataFrame()), "gap_flag", "✓ OK")
    _write_sheet(writer, "📒 Ledger Receipts",    results.get("ledger_by_sku",   pd.DataFrame()))
    _write_sheet(writer, "📋 Ledger Detail",      results.get("ledger_receipts", pd.DataFrame()))
    _write_sheet(writer, "⏳ Currently Inbound",  results.get("currently_inbound", pd.DataFrame()))

    writer.close()
    print(f"\n  ✓ Saved: {output_path}")


# ─── Main ──────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description="FBA Inbound Gap Checker")
    parser.add_argument("--days",       type=int, default=90,  help="Lookback days (default 90)")
    parser.add_argument("--no-ledger",  action="store_true",   help="Skip inventory ledger (faster, no receipt cross-check)")
    args = parser.parse_args()

    end_dt   = datetime.now(timezone.utc).replace(hour=0, minute=0, second=0, microsecond=0) - timedelta(seconds=1)
    start_dt = (end_dt - timedelta(days=args.days - 1)).replace(hour=0, minute=0, second=0, microsecond=0)

    period_label = f"Last {args.days} days (up to {end_dt.date()})"
    print(f"\n{'='*65}")
    print(f"  FBA Inbound Gap Checker")
    print(f"  Period: {start_dt.date()} → {end_dt.date()} ({args.days} days)")
    print(f"{'='*65}")

    # 1. Inbound shipments (DATE_RANGE method — no ShipmentStatusList needed)
    shipments      = fetch_inbound_shipments_date_range(start_dt)
    shipment_items = pd.DataFrame()
    if shipments:
        shipment_items = fetch_shipment_items(shipments)
    else:
        print("  ⚠ No shipments returned — inbound comparison not possible via API")

    # 2. Inventory ledger Receipts
    ledger_receipts = pd.DataFrame()
    if not args.no_ledger:
        try:
            ledger_receipts, _ = fetch_ledger_receipts(start_dt, end_dt)
        except Exception as e:
            print(f"  ⚠ Ledger failed: {e}")
    else:
        print("\n  [3] Ledger skipped (--no-ledger)")

    # 3. Inventory snapshot (current inbound qty)
    inventory_snapshot = pd.DataFrame()
    try:
        inventory_snapshot = fetch_inventory_snapshot()
    except Exception as e:
        print(f"  ⚠ Inventory snapshot failed: {e}")

    # 4. Analyse
    print(f"\n{'='*65}")
    print(f"  ANALYSIS")
    print(f"{'='*65}")
    results = analyse_gaps(shipment_items, ledger_receipts, inventory_snapshot)

    # 5. Save Excel
    ts          = datetime.now().strftime("%Y%m%d_%H%M")
    output_path = os.path.join(OUTPUT_DIR, f"inbound_gaps_{args.days}d_{ts}.xlsx")
    write_excel(results, output_path, period_label)
    print(f"\n  Done.")


if __name__ == "__main__":
    main()
