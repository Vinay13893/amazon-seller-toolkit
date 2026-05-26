"""
FBA Reconciliation — Excel Report Builder
==========================================
Takes all analysis DataFrames and writes a single colour-coded Excel
workbook with multiple sheets.
"""

import os
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

# ─── Colour palette ───────────────────────────────────────────
CLR_HEADER_BG   = "1F4E79"  # dark blue
CLR_HEADER_FG   = "FFFFFF"
CLR_SECTION_BG  = "D6E4F0"
CLR_OK          = "C6EFCE"  # green
CLR_WARN        = "FFEB9C"  # amber
CLR_ALERT       = "FFC7CE"  # red
CLR_NEUTRAL     = "EBF3FB"  # very light blue (alt row)
CLR_WHITE       = "FFFFFF"

THIN = Side(style="thin", color="B0B0B0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _header_fill(hex_color=CLR_HEADER_BG):
    return PatternFill("solid", fgColor=hex_color)


def _cell_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)


def _style_header_row(ws, row_num, n_cols):
    for col in range(1, n_cols + 1):
        cell            = ws.cell(row=row_num, column=col)
        cell.font       = Font(name="Calibri", bold=True, color=CLR_HEADER_FG, size=10)
        cell.fill       = _header_fill()
        cell.alignment  = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border     = BORDER


def _auto_width(ws, min_w=10, max_w=50):
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = max(min_w, min(max_len + 2, max_w))


def _write_df_to_sheet(ws, df, title=None, freeze_row=2, alt_rows=True):
    start_row = 1
    if title:
        ws.cell(row=1, column=1, value=title).font = Font(
            bold=True, size=12, color="1F4E79"
        )
        ws.cell(row=1, column=1).fill = _cell_fill(CLR_SECTION_BG)
        ws.merge_cells(
            start_row=1, start_column=1,
            end_row=1,   end_column=max(len(df.columns), 1)
        )
        start_row = 2

    if df.empty:
        ws.cell(row=start_row, column=1, value="No data available for this period.")
        return

    # Headers
    for ci, col_name in enumerate(df.columns, 1):
        ws.cell(row=start_row, column=ci, value=col_name)
    _style_header_row(ws, start_row, len(df.columns))

    # Data rows
    for ri, (_, row) in enumerate(df.iterrows(), start_row + 1):
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font      = Font(name="Calibri", size=10)
            cell.border    = BORDER
            cell.alignment = Alignment(horizontal="left", vertical="center")
            if alt_rows and (ri % 2 == 0):
                cell.fill = _cell_fill(CLR_NEUTRAL)
            else:
                cell.fill = _cell_fill(CLR_WHITE)

    ws.freeze_panes = ws.cell(row=freeze_row + (1 if title else 0), column=1)
    _auto_width(ws)


def _colour_reconciliation_sheet(ws, df, title_offset):
    """Colour-code variance and status columns in the reconciliation sheet."""
    if df.empty:
        return
    cols = list(df.columns)
    status_ci  = cols.index("status")  + 1 if "status"   in cols else None
    var_ci     = cols.index("variance") + 1 if "variance" in cols else None
    unf_ci     = cols.index("unfulfillable") + 1 if "unfulfillable" in cols else None

    data_start = title_offset + 2   # title row + header row
    for ri in range(data_start, ws.max_row + 1):
        status_val = ws.cell(row=ri, column=status_ci).value if status_ci else None
        var_val    = ws.cell(row=ri, column=var_ci).value    if var_ci    else None
        unf_val    = ws.cell(row=ri, column=unf_ci).value    if unf_ci    else None

        row_color = None
        if isinstance(var_val, (int, float)):
            if var_val == 0:    row_color = CLR_OK
            elif var_val < 0:   row_color = CLR_ALERT
            elif var_val > 0:   row_color = CLR_WARN

        if row_color:
            for ci in range(1, ws.max_column + 1):
                ws.cell(row=ri, column=ci).fill = _cell_fill(row_color)

        # Extra red on unfulfillable > 0
        if unf_val and isinstance(unf_val, int) and unf_val > 0:
            ws.cell(row=ri, column=unf_ci).fill = _cell_fill(CLR_ALERT)
            ws.cell(row=ri, column=unf_ci).font = Font(bold=True, color="9C0006")


def _colour_returns_sheet(ws, df, title_offset):
    if df.empty:
        return
    cols = list(df.columns)
    for col_name, color in [
        ("returned_sellable",   CLR_OK),
        ("customer_damaged",    CLR_ALERT),
        ("defective",           CLR_ALERT),
        ("warehouse_damaged",   CLR_WARN),
        ("carrier_damaged",     CLR_WARN),
        ("expired",             CLR_ALERT),
        ("total_unsellable",    CLR_ALERT),
    ]:
        if col_name in cols:
            ci = cols.index(col_name) + 1
            data_start = title_offset + 2
            for ri in range(data_start, ws.max_row + 1):
                val = ws.cell(row=ri, column=ci).value
                if isinstance(val, int) and val > 0:
                    ws.cell(row=ri, column=ci).fill = _cell_fill(color)


def _colour_inbound_sheet(ws, df, title_offset):
    if df.empty:
        return
    cols = list(df.columns)
    disc_ci = cols.index("discrepancy") + 1 if "discrepancy" in cols else None
    if not disc_ci:
        return
    data_start = title_offset + 2
    for ri in range(data_start, ws.max_row + 1):
        val = ws.cell(row=ri, column=disc_ci).value
        if isinstance(val, (int, float)) and val < 0:
            for ci in range(1, ws.max_column + 1):
                ws.cell(row=ri, column=ci).fill = _cell_fill(CLR_ALERT)


# ─── Main Excel Builder ───────────────────────────────────────
def build_excel(
    output_path: str,
    recon_df,
    returns_detail_df,
    returns_total_df,
    unfulfillable_df,
    reimb_detail_df,
    reimb_total_df,
    inbound_df,
    period_label: str,
):
    print(f"\n  Building Excel report → {output_path}")
    writer = pd.ExcelWriter(output_path, engine="openpyxl")

    # ── Sheet 1: Dashboard / Summary ──────────────────────────
    wb_dummy = writer.book

    def _strip_tz(df):
        """Remove timezone from all datetime columns so Excel doesn't complain."""
        if df is None or df.empty:
            return df
        df = df.copy()
        for col in df.columns:
            if pd.api.types.is_datetime64_any_dtype(df[col]):
                try:
                    df[col] = df[col].dt.tz_localize(None)
                except TypeError:
                    df[col] = df[col].dt.tz_convert(None)
        return df

    sheets = [
        ("📊 Reconciliation",  recon_df),
        ("📦 Returns Detail",  returns_total_df),
        ("🔴 Unfulfillable",   unfulfillable_df),
        ("💰 Reimbursements",  reimb_total_df),
        ("🚚 Inbound Gaps",    inbound_df),
        ("📋 Returns Full",    returns_detail_df),
        ("💳 Reimb Detail",    reimb_detail_df),
    ]

    for sheet_name, df in sheets:
        df = _strip_tz(df)
        if df is None:
            df = pd.DataFrame()
        df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=1)
        ws = writer.sheets[sheet_name]
        # Write title banner
        ws.cell(row=1, column=1, value=f"FBA Reconciliation — {period_label}   |   {sheet_name}")
        ws.cell(row=1, column=1).font = Font(bold=True, size=12, color=CLR_HEADER_BG)
        ws.cell(row=1, column=1).fill = _cell_fill(CLR_SECTION_BG)
        if not df.empty:
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))
        # Style header row (row 2 because of the title row)
        if not df.empty:
            _style_header_row(ws, 2, len(df.columns))
        _auto_width(ws)
        ws.freeze_panes = ws.cell(row=3, column=1)

    writer.close()

    # ── Post-process: colour-code using openpyxl ──────────────
    wb = load_workbook(output_path)

    if "📊 Reconciliation" in wb.sheetnames and recon_df is not None and not recon_df.empty:
        _colour_reconciliation_sheet(wb["📊 Reconciliation"], recon_df, title_offset=1)

    if "📦 Returns Detail" in wb.sheetnames and returns_total_df is not None and not returns_total_df.empty:
        _colour_returns_sheet(wb["📦 Returns Detail"], returns_total_df, title_offset=1)

    if "🚚 Inbound Gaps" in wb.sheetnames and inbound_df is not None and not inbound_df.empty:
        _colour_inbound_sheet(wb["🚚 Inbound Gaps"], inbound_df, title_offset=1)

    # Dashboard sheet
    _build_dashboard(wb, recon_df, returns_total_df, unfulfillable_df,
                     reimb_total_df, inbound_df, period_label)

    wb.save(output_path)
    print(f"  ✓ Saved: {output_path}")


def _build_dashboard(wb, recon_df, returns_df, unfulfillable_df,
                     reimb_df, inbound_df, period_label):
    ws = wb.create_sheet("🏠 Dashboard", 0)

    def _kv(row, label, value, color=None):
        label_cell = ws.cell(row=row, column=1, value=label)
        label_cell.font = Font(bold=True, name="Calibri", size=11)
        label_cell.fill = _cell_fill(CLR_SECTION_BG)
        label_cell.alignment = Alignment(horizontal="left")
        val_cell = ws.cell(row=row, column=2, value=value)
        val_cell.font = Font(name="Calibri", size=11, bold=bool(color))
        if color:
            val_cell.fill = _cell_fill(color)
        val_cell.alignment = Alignment(horizontal="left")

    # Title
    ws.cell(row=1, column=1,
            value=f"FBA Inventory Reconciliation — {period_label}").font = Font(
        bold=True, size=14, color=CLR_HEADER_BG)
    ws.cell(row=1, column=1).fill = _cell_fill(CLR_SECTION_BG)
    ws.merge_cells("A1:E1")

    ws.cell(row=2, column=1,
            value=f"Generated: {datetime.now().strftime('%d %b %Y %H:%M')}").font = Font(
        size=10, color="888888")

    row = 4
    ws.cell(row=row, column=1, value="METRIC").font   = Font(bold=True)
    ws.cell(row=row, column=2, value="VALUE").font    = Font(bold=True)
    ws.cell(row=row, column=3, value="ACTION").font   = Font(bold=True)
    for ci in [1,2,3]:
        ws.cell(row=row, column=ci).fill = _header_fill()
        ws.cell(row=row, column=ci).font = Font(bold=True, color="FFFFFF")
    row += 1

    # ── Reconciliation summary
    if recon_df is not None and not recon_df.empty:
        total_sold  = int(recon_df["units_sold"].sum()) if "units_sold" in recon_df else 0
        total_recv  = int(recon_df["units_received"].sum()) if "units_received" in recon_df else 0
        total_avail = int(recon_df["actual_available"].sum()) if "actual_available" in recon_df else 0
        under_count = int((recon_df["variance"] < 0).sum()) if "variance" in recon_df else 0
        over_count  = int((recon_df["variance"] > 0).sum()) if "variance" in recon_df else 0

        _kv(row, "Total Units Sold (period)",     total_sold);    row += 1
        _kv(row, "Total Units Received (inbound)", total_recv);   row += 1
        _kv(row, "Total Available Stock (now)",   total_avail);   row += 1
        _kv(row, "SKUs with UNDER variance",      under_count,
            CLR_ALERT if under_count > 0 else CLR_OK);            row += 1
        _kv(row, "SKUs with OVER variance",       over_count,
            CLR_WARN if over_count > 0 else CLR_OK);              row += 1

    row += 1
    # ── Unfulfillable
    if unfulfillable_df is not None and not unfulfillable_df.empty:
        total_unf = int(unfulfillable_df["unfulfillable_qty"].sum()) if "unfulfillable_qty" in unfulfillable_df else 0
        _kv(row, "Total Unfulfillable Units",     total_unf,
            CLR_ALERT if total_unf > 0 else CLR_OK)
        ws.cell(row=row, column=3, value="→ Place Removal / Disposal Order in Seller Central")
        ws.cell(row=row, column=3).font = Font(italic=True, color="9C0006")
        row += 1

    row += 1
    # ── Returns
    if returns_df is not None and not returns_df.empty:
        total_ret  = int(returns_df["total_returns"].sum())    if "total_returns"    in returns_df else 0
        total_sell = int(returns_df["returned_sellable"].sum()) if "returned_sellable" in returns_df else 0
        total_dam  = int(returns_df["total_unsellable"].sum()) if "total_unsellable" in returns_df else 0

        _kv(row, "Total Returns (period)",        total_ret);    row += 1
        _kv(row, "  → Returned Sellable (back in stock)", total_sell, CLR_OK); row += 1
        _kv(row, "  → Returned Unsellable (stuck / lost)", total_dam,
            CLR_ALERT if total_dam > 0 else CLR_OK);             row += 1

    row += 1
    # ── Reimbursements
    if reimb_df is not None and not reimb_df.empty:
        total_amount = reimb_df["total_amount"].sum() if "total_amount" in reimb_df else 0
        total_units  = int(reimb_df["total_qty"].sum()) if "total_qty" in reimb_df else 0
        _kv(row, "Total Reimbursed (period)",
            f"₹{total_amount:,.0f}  ({total_units} units)");    row += 1

    row += 1
    # ── Inbound discrepancies
    if inbound_df is not None and not inbound_df.empty:
        short_units = int(inbound_df["discrepancy"].sum()) if "discrepancy" in inbound_df else 0
        _kv(row, "Short-Received Units (inbound)", abs(short_units),
            CLR_ALERT if short_units < 0 else CLR_OK)
        ws.cell(row=row, column=3,
                value="→ File reimbursement claim for CLOSED shipments with discrepancy")
        ws.cell(row=row, column=3).font = Font(italic=True, color="9C0006")
        row += 1

    row += 2
    ws.cell(row=row, column=1,
            value="📌 How to read this report").font = Font(bold=True, size=11, color=CLR_HEADER_BG)
    row += 1
    notes = [
        ("📊 Reconciliation",  "Month-end formula per SKU. GREEN = balanced, RED = Amazon owes you, AMBER = over-reported."),
        ("📦 Returns Detail",  "All returns broken down by disposition. Focus on Customer Damaged & Defective."),
        ("🔴 Unfulfillable",   "Stock Amazon can't sell. Needs Removal or Disposal order."),
        ("💰 Reimbursements",  "All cash/unit reimbursements Amazon paid you. Cross-check with Unfulfillable."),
        ("🚚 Inbound Gaps",    "Shipments where you sent more than Amazon received. File claims for CLOSED shipments."),
        ("📋 Returns Full",    "Every individual return transaction with reason and disposition."),
        ("💳 Reimb Detail",    "Every individual reimbursement event."),
    ]
    for sheet, note in notes:
        ws.cell(row=row, column=1, value=sheet).font = Font(bold=True, size=10)
        ws.cell(row=row, column=2, value=note).font  = Font(size=10)
        ws.cell(row=row, column=2).alignment = Alignment(wrap_text=True)
        row += 1

    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 60
    ws.row_dimensions[1].height     = 30
