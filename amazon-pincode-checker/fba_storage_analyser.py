"""
FBA Storage Cost Analyser — Ledger Edition
==========================================
Reads: Emount Amazon FBA Stock Volume 11052026.xlsx  (per-unit volume data)
Pulls: 30-day Inventory Ledger via SP-API             (actual sales / inward / movements)
Outputs: fba_storage_analysis_<date>.xlsx

What it analyses:
  1. Dashboard             — total KPIs: rent, stock, velocity
  2. Stock Movement        — opening → receipts → sales → returns → closing per SKU
  3. Sales Velocity        — 30d units sold, daily rate, days-of-stock remaining
  4. Replenishment Alerts  — stock that will run out soon based on actual velocity
  5. Dead Stock            — zero sales SKUs burning rent (remove / liquidate)
  6. Rent Ranking          — top rent burners with monthly projection
  7. Rent Per Unit         — cost per unit per month (holding cost per item)
  8. FC Spread             — SKUs in too many FCs
  9. Full Ledger           — every event for 30 days
"""

import os
import sys
import time
import pandas as pd
from datetime import datetime, timezone, timedelta

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
SOURCE_FILE = os.path.join(SCRIPT_DIR, "Emount Amazon FBA Stock Volume 11052026.xlsx")
OUTPUT_DIR  = os.path.join(SCRIPT_DIR, "output")
os.makedirs(OUTPUT_DIR, exist_ok=True)

sys.path.insert(0, os.path.join(SCRIPT_DIR, "fba_reconciliation"))
from sp_client import _run_report

RATE_PER_M3_PER_DAY = 60.0   # Amazon India FBA storage rate
DAYS = 30                     # ledger lookback window



# ─── Step 1: Load volume file (per-unit m3 data) ──────────────
def load_volume_file():
    """
    Reads Only FBA sheet for per-item volume.
    Returns per-SKU: per_item_volume, current_ending_units, FCs.
    """
    xl  = pd.ExcelFile(SOURCE_FILE, engine="openpyxl")
    fba = xl.parse("Only FBA")
    for col in ["Ending Warehouse Balance", "Per Item Volume", "Total Volume Used"]:
        if col in fba.columns:
            fba[col] = pd.to_numeric(fba[col], errors="coerce").fillna(0)

    # Per-SKU volume (average across FCs — same product, same volume)
    vol = (
        fba.groupby("MSKU")
        .agg(
            per_item_volume  = ("Per Item Volume",          "mean"),
            current_units    = ("Ending Warehouse Balance", "sum"),
            total_volume_m3  = ("Total Volume Used",        "sum"),
            fc_count         = ("Location",                 "nunique"),
            fcs              = ("Location",                 lambda x: ", ".join(sorted(x[fba.loc[x.index, "Ending Warehouse Balance"] > 0].dropna().unique()))),
        )
        .reset_index()
        .rename(columns={"MSKU": "sku"})
    )
    vol["daily_rent_now"]   = (vol["total_volume_m3"] * RATE_PER_M3_PER_DAY).round(4)
    vol["monthly_rent_now"] = (vol["daily_rent_now"]  * 30).round(2)

    print(f"  Volume file: {len(vol)} SKUs, {int(vol['current_units'].sum())} units, "
          f"{vol['total_volume_m3'].sum():.3f} m3")
    print(f"  Current monthly rent: Rs {vol['monthly_rent_now'].sum():,.2f}")
    return vol, fba


# ─── Step 2: Fetch 30-day inventory ledger ────────────────────
def fetch_ledger(start_dt, end_dt):
    print(f"\n  Fetching ledger: {start_dt.date()} → {end_dt.date()} ...")
    df = _run_report("GET_LEDGER_DETAIL_VIEW_DATA", start_dt, end_dt)
    df["quantity"] = pd.to_numeric(df["quantity"], errors="coerce").fillna(0).astype(int)
    print(f"  Ledger rows: {len(df)}")
    print(f"  Event types:\n{df['event-type'].value_counts().to_string()}")
    return df


# ─── Step 3: Build stock movement summary ─────────────────────
def build_stock_movement(ledger_df, vol_df):
    """
    Per-SKU 30-day movement from ledger:
      Receipts      = inward (positive, event-type = Receipts)
      Shipments     = sales  (negative, event-type = Shipments)
      CustomerRet   = returns back (positive, event-type = CustomerReturns)
      WhseTransfers = FC-to-FC moves (net should be 0 for whole account)
      VendorReturns = removals (negative, event-type = VendorReturns)
      Adjustments   = found/lost/damaged
    """
    grp = ledger_df.groupby(["msku", "event-type"])["quantity"].sum().reset_index()

    # Pivot event types into columns
    pivot = grp.pivot_table(index="msku", columns="event-type", values="quantity",
                            aggfunc="sum", fill_value=0).reset_index()
    pivot.columns.name = None
    pivot = pivot.rename(columns={"msku": "sku"})

    # Normalise expected columns
    for col in ["Receipts", "Shipments", "CustomerReturns", "WhseTransfers",
                "VendorReturns", "Adjustments", "MisplacedAndFoundAdjustments"]:
        if col not in pivot.columns:
            pivot[col] = 0

    pivot["units_sold"]       = pivot["Shipments"].abs()          # negative = sold
    pivot["units_received"]   = pivot["Receipts"]                 # positive = inward
    pivot["units_returned"]   = pivot["CustomerReturns"]          # positive = came back
    pivot["units_removed"]    = pivot["VendorReturns"].abs()      # negative = removed
    pivot["net_adjustments"]  = (
        pivot.get("Adjustments", 0) +
        pivot.get("MisplacedAndFoundAdjustments", 0)
    )

    # Net change = received + returned - sold - removed + adjustments
    pivot["net_change"] = (
        pivot["units_received"] + pivot["units_returned"]
        - pivot["units_sold"] - pivot["units_removed"]
        + pivot["net_adjustments"]
    )

    # Merge current stock from volume file
    merged = pivot.merge(
        vol_df[["sku", "per_item_volume", "current_units",
                "total_volume_m3", "daily_rent_now", "monthly_rent_now",
                "fc_count", "fcs"]],
        on="sku", how="outer"
    ).fillna(0)

    # Daily sales velocity and days of stock remaining
    merged["daily_velocity"]   = (merged["units_sold"] / DAYS).round(2)
    merged["days_stock_left"]  = (
        merged["current_units"] / merged["daily_velocity"].replace(0, 0.001)
    ).round(0).astype(int)
    merged["weeks_stock_left"] = (merged["days_stock_left"] / 7).round(1)

    # Replenishment flag
    merged["stock_flag"] = merged.apply(
        lambda r: (
            "🔴 DEAD STOCK — REMOVE"        if r["units_sold"] == 0 and r["current_units"] > 0
            else "🔴 OUT OF STOCK"           if r["current_units"] <= 0
            else "🟠 REORDER NOW (<2 weeks)" if r["days_stock_left"] < 14
            else "🟡 REORDER SOON (<4 wks)"  if r["days_stock_left"] < 28
            else "🟠 OVERSTOCKED (>90 days)" if r["days_stock_left"] > 90
            else "🟢 OK"
        ), axis=1
    )

    return merged


# ─── Step 4: Derived analysis sheets ─────────────────────────
def build_sales_velocity(movement_df):
    cols = ["sku", "units_sold", "daily_velocity", "current_units",
            "days_stock_left", "weeks_stock_left", "monthly_rent_now",
            "units_received", "units_returned", "units_removed", "stock_flag"]
    df = movement_df[cols].copy()
    df = df.sort_values("units_sold", ascending=False).reset_index(drop=True)
    df.insert(0, "rank", range(1, len(df) + 1))
    return df


def build_replenishment_alerts(movement_df):
    alert = movement_df[
        (movement_df["days_stock_left"] < 28) &
        (movement_df["units_sold"] > 0) &
        (movement_df["current_units"] > 0)
    ].copy()
    cols = ["sku", "current_units", "days_stock_left", "weeks_stock_left",
            "daily_velocity", "units_sold", "monthly_rent_now", "stock_flag"]
    alert = alert[cols].sort_values("days_stock_left").reset_index(drop=True)
    print(f"\n  Replenishment alerts: {len(alert)} SKUs running low")
    for _, r in alert.iterrows():
        print(f"    {r['sku']:<45} {int(r['current_units'])} units  "
              f"{int(r['days_stock_left'])} days left  "
              f"({r['daily_velocity']:.1f}/day)  {r['stock_flag']}")
    return alert


def build_dead_stock(movement_df):
    dead = movement_df[
        (movement_df["units_sold"] == 0) &
        (movement_df["current_units"] > 0)
    ].copy()
    cols = ["sku", "current_units", "monthly_rent_now", "fc_count",
            "units_received", "units_removed", "net_adjustments", "stock_flag"]
    dead = dead[cols].sort_values("monthly_rent_now", ascending=False).reset_index(drop=True)
    total_burn = dead["monthly_rent_now"].sum()
    print(f"\n  Dead stock: {len(dead)} SKUs, "
          f"{int(dead['current_units'].sum())} units, "
          f"Rs {total_burn:,.2f}/month burn")
    return dead


def build_rent_ranking(movement_df):
    df = movement_df[movement_df["current_units"] > 0].copy()
    df["pct_of_total"] = (df["monthly_rent_now"] / df["monthly_rent_now"].sum() * 100).round(1)
    df["cumulative_pct"] = df.sort_values("monthly_rent_now", ascending=False)["pct_of_total"].cumsum().round(1)
    cols = ["sku", "current_units", "monthly_rent_now", "per_item_volume",
            "units_sold", "daily_velocity", "days_stock_left",
            "pct_of_total", "cumulative_pct", "stock_flag"]
    df = df[cols].sort_values("monthly_rent_now", ascending=False).reset_index(drop=True)
    df.insert(0, "rank", range(1, len(df) + 1))

    print(f"\n  Top 10 rent burners (30-day ledger):")
    for _, r in df.head(10).iterrows():
        print(f"    {int(r['rank']):2}. {r['sku']:<45} "
              f"Rs {r['monthly_rent_now']:7,.2f}/mo  "
              f"{int(r['current_units'])} units  "
              f"{int(r['units_sold'])} sold/30d  "
              f"{r['pct_of_total']}%")
    return df


def build_fc_spread(vol_df):
    spread = vol_df[vol_df["current_units"] > 0].copy()
    spread["flag"] = spread["fc_count"].apply(
        lambda n: "🔴 HIGH SPREAD (>8)" if n > 8
        else      "🟠 WIDE (5-8)"        if n >= 5
        else      "🟡 MODERATE (3-4)"    if n >= 3
        else      "🟢 OK (1-2)"
    )
    spread = spread[["sku", "current_units", "fc_count", "fcs",
                     "monthly_rent_now", "flag"]].sort_values("fc_count", ascending=False).reset_index(drop=True)
    return spread


# ─── Step 5: Dashboard ────────────────────────────────────────
def build_dashboard(movement_df, start_dt, end_dt):
    total_units    = int(movement_df["current_units"].sum())
    total_monthly  = movement_df["monthly_rent_now"].sum()
    total_sold_30d = int(movement_df["units_sold"].sum())
    total_received = int(movement_df["units_received"].sum())
    dead_mask      = (movement_df["units_sold"] == 0) & (movement_df["current_units"] > 0)
    dead_count     = int(dead_mask.sum())
    dead_units     = int(movement_df.loc[dead_mask, "current_units"].sum())
    dead_rent      = movement_df.loc[dead_mask, "monthly_rent_now"].sum()
    reorder_mask   = (movement_df["days_stock_left"] < 14) & (movement_df["units_sold"] > 0) & (movement_df["current_units"] > 0)
    reorder_count  = int(reorder_mask.sum())

    top_rent = movement_df.sort_values("monthly_rent_now", ascending=False).iloc[0]

    rows = [
        ["", ""],
        ["PERIOD",                    f"{start_dt.date()} → {end_dt.date()} ({DAYS} days)"],
        ["", ""],
        ["TOTAL FBA UNITS (now)",     f"{total_units:,}"],
        ["TOTAL MONTHLY RENT",        f"Rs {total_monthly:,.2f}"],
        ["RATE USED",                 f"Rs {RATE_PER_M3_PER_DAY}/m3/day"],
        ["", ""],
        ["UNITS SOLD (30 days)",      f"{total_sold_30d:,}"],
        ["UNITS RECEIVED (30 days)",  f"{total_received:,}"],
        ["", ""],
        ["BIGGEST RENT BURNER",       top_rent["sku"]],
        ["  Monthly cost",            f"Rs {top_rent['monthly_rent_now']:,.2f}  ({top_rent['units_sold']} sold / {int(top_rent['current_units'])} in stock)"],
        ["", ""],
        ["DEAD STOCK (0 sales)",      f"{dead_count} SKUs — {dead_units} units"],
        ["Dead stock rent burn",      f"Rs {dead_rent:,.2f}/month — RECOVERABLE by removal"],
        ["", ""],
        ["REORDER ALERTS (<2 weeks)", f"{reorder_count} SKUs running low"],
        ["", ""],
        ["ANALYSIS DATE",             datetime.now().strftime("%Y-%m-%d %H:%M")],
        ["SOURCE FILE",               os.path.basename(SOURCE_FILE)],
    ]
    df = pd.DataFrame(rows, columns=["Metric", "Value"])

    print(f"\n{'='*65}")
    print(f"  DASHBOARD")
    print(f"{'='*65}")
    for k, v in rows:
        if k:
            print(f"  {k:<30} {v}")
    print(f"{'='*65}")
    return df, total_monthly


# ─── Step 6: Excel writer ─────────────────────────────────────
def write_excel(output_path, dashboard_df, total_monthly, velocity_df,
                reorder_df, dead_df, ranking_df, fc_spread_df, ledger_df):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    CLR_HDR   = "1F4E79"
    CLR_OK    = "C6EFCE"
    CLR_WARN  = "FFEB9C"
    CLR_ALERT = "FFC7CE"
    CLR_DASH  = "2E75B6"
    THIN = Side(style="thin", color="B0B0B0")
    BRD  = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    def _strip_tz(df):
        df = df.copy()
        for col in df.columns:
            if pd.api.types.is_datetime64_any_dtype(df[col]):
                try:    df[col] = df[col].dt.tz_localize(None)
                except: df[col] = df[col].dt.tz_convert(None)
        return df

    def _write_sheet(writer, name, df, flag_col=None):
        if df is None or df.empty:
            df = pd.DataFrame({"note": ["No data"]})
        df = _strip_tz(df)
        df.to_excel(writer, sheet_name=name, index=False, startrow=1)
        ws = writer.sheets[name]

        title = f"FBA Storage Analysis  |  {name}  |  Monthly rent: Rs {total_monthly:,.0f}"
        ws.cell(row=1, column=1, value=title).font = Font(bold=True, size=12, color="FFFFFF")
        ws.cell(row=1, column=1).fill = PatternFill("solid", fgColor=CLR_HDR)
        if len(df.columns) > 1:
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))

        for ci in range(1, len(df.columns) + 1):
            c = ws.cell(row=2, column=ci)
            c.font      = Font(bold=True, color="FFFFFF", size=10)
            c.fill      = PatternFill("solid", fgColor=CLR_HDR)
            c.alignment = Alignment(horizontal="center", wrap_text=True)
            c.border    = BRD

        for ri in range(3, ws.max_row + 1):
            row_color = None
            if flag_col and flag_col in df.columns:
                ci_f = list(df.columns).index(flag_col) + 1
                val  = str(ws.cell(row=ri, column=ci_f).value or "")
                row_color = (
                    CLR_ALERT if "🔴" in val
                    else CLR_WARN if "🟠" in val or "🟡" in val
                    else CLR_OK   if "🟢" in val else None
                )
            for ci in range(1, len(df.columns) + 1):
                c = ws.cell(row=ri, column=ci)
                c.font      = Font(size=10)
                c.alignment = Alignment(horizontal="left")
                c.border    = BRD
                if row_color:
                    c.fill = PatternFill("solid", fgColor=row_color)

        for col_cells in ws.columns:
            w = max((len(str(c.value or "")) for c in col_cells), default=8)
            ws.column_dimensions[get_column_letter(col_cells[0].column)].width = max(12, min(w + 2, 65))
        ws.freeze_panes = ws.cell(row=3, column=1)

    def _write_dashboard(writer, df):
        df.to_excel(writer, sheet_name="🏠 Dashboard", index=False, startrow=1)
        ws = writer.sheets["🏠 Dashboard"]
        ws.cell(row=1, column=1, value="FBA Storage Analysis — Dashboard").font = Font(bold=True, size=14, color="FFFFFF")
        ws.cell(row=1, column=1).fill = PatternFill("solid", fgColor=CLR_DASH)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
        for ri in range(2, ws.max_row + 1):
            k = ws.cell(row=ri, column=1)
            v = ws.cell(row=ri, column=2)
            k_val = str(k.value or "")
            k.font = Font(bold=True, size=11, color=CLR_HDR if k_val else "FFFFFF")
            v.font = Font(size=11)
            k.border = v.border = BRD
            if "MONTHLY RENT" in k_val:
                v.font = Font(bold=True, size=13, color="C00000")
            if "DEAD" in k_val or "burn" in k_val.lower():
                v.font = Font(bold=True, color="C00000")
            if "REORDER" in k_val:
                v.font = Font(bold=True, color="FF6600")
        ws.column_dimensions["A"].width = 34
        ws.column_dimensions["B"].width = 60

    writer = pd.ExcelWriter(output_path, engine="openpyxl")
    _write_dashboard(writer, dashboard_df)
    _write_sheet(writer, "📦 Stock Movement",     velocity_df,  "stock_flag")
    _write_sheet(writer, "🔁 Replenishment",       reorder_df,   "stock_flag")
    _write_sheet(writer, "🔴 Dead Stock",          dead_df,      "stock_flag")
    _write_sheet(writer, "📊 Rent Ranking",        ranking_df,   "stock_flag")
    _write_sheet(writer, "🏢 FC Spread",           fc_spread_df, "flag")
    _write_sheet(writer, "📋 Full Ledger",         ledger_df)
    writer.close()
    print(f"\n  ✓ Saved: {output_path}")


# ─── Main ─────────────────────────────────────────────────────
def main():
    print(f"\n{'='*65}")
    print(f"  FBA Storage Cost Analyser — Ledger Edition")
    print(f"{'='*65}")

    end_dt   = datetime.now(timezone.utc).replace(hour=0, minute=0, second=0, microsecond=0) - timedelta(seconds=1)
    start_dt = (end_dt - timedelta(days=DAYS - 1)).replace(hour=0, minute=0, second=0, microsecond=0)

    # Load per-unit volume
    vol_df, _ = load_volume_file()

    # Fetch 30-day ledger
    ledger_df = fetch_ledger(start_dt, end_dt)

    # Build movement summary
    movement_df = build_stock_movement(ledger_df, vol_df)

    # Derived sheets
    velocity_df = build_sales_velocity(movement_df)
    reorder_df  = build_replenishment_alerts(movement_df)
    dead_df     = build_dead_stock(movement_df)
    ranking_df  = build_rent_ranking(movement_df)
    fc_df       = build_fc_spread(vol_df)
    dash_df, total_monthly = build_dashboard(movement_df, start_dt, end_dt)

    # Write Excel
    ts          = datetime.now().strftime("%Y%m%d_%H%M")
    output_path = os.path.join(OUTPUT_DIR, f"fba_storage_analysis_{ts}.xlsx")
    write_excel(output_path, dash_df, total_monthly, velocity_df,
                reorder_df, dead_df, ranking_df, fc_df, ledger_df)
    print(f"\n  Done.")


if __name__ == "__main__":
    main()
